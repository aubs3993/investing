"""Scaffold templates/broker_fetcher.xlsx — live broker-estimate workbook.

Idempotent. Same conventions as scaffold_capiq_fetcher.py.

Layout MUST mirror _Broker_Data on the main template. Both read from
shared.broker_layout to keep them in lockstep. Column A is a 2.71-wide
spacer; labels live in column B and primary values/inputs in column C.

Run:
    python -m shared.scaffold_broker_fetcher
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName

from shared import broker_layout
# Styling constants + style_* helpers are shared across all scaffolders.
from shared.excel_helpers import (
    BANNER_FONT, LABEL, LABEL_BOLD, TITLE_FONT,
    style_formula, style_header, style_input,
)

REPO_ROOT = Path(__file__).resolve().parent.parent
FETCHER_PATH = REPO_ROOT / "templates" / "broker_fetcher.xlsx"
TICKER_REF = "broker_fetcher_ticker"


def _est_args(ticker_ref, period, omit_currency):
    if omit_currency:
        return f"{ticker_ref},{period}"
    return f"{ticker_ref},{period},IQ_USD"


def build_fetcher(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2.71

    title_cell = ws.cell(broker_layout.ROW_TITLE, 2,
                         f'={TICKER_REF}&" | Broker Estimates Fetcher"')
    title_cell.font = TITLE_FONT

    ws.cell(broker_layout.ROW_BANNER, 2,
            "Broker Estimates Fetcher — formulas only. Open in Excel with CapIQ plugin loaded to refresh."
            ).font = BANNER_FONT

    ws.cell(broker_layout.ROW_LAST_FETCH, 2, "Run via:").font = LABEL_BOLD
    ws.cell(broker_layout.ROW_LAST_FETCH, 3,
            "python -m shared.fetch_broker_estimates <TICKER>").font = LABEL

    ws.cell(broker_layout.ROW_TICKER, 2, "Ticker:").font = LABEL_BOLD
    ticker_cell = ws.cell(broker_layout.ROW_TICKER, 3, "AAPL")
    style_input(ticker_cell)
    ticker_cell.comment = Comment(
        "Set this to the ticker you want broker estimates for. CapIQ formulas "
        "refresh automatically when you change it (and when fetch_broker_estimates.py "
        "drives it programmatically).",
        "scaffold_broker_fetcher",
    )

    ws.cell(broker_layout.ROW_FY1_YEAR, 2, "FY1 fiscal year:").font = LABEL_BOLD
    ws.cell(broker_layout.ROW_FY2_YEAR, 2, "FY2 fiscal year:").font = LABEL_BOLD
    ws.cell(broker_layout.ROW_FY3_YEAR, 2, "FY3 fiscal year:").font = LABEL_BOLD
    style_formula(ws.cell(broker_layout.ROW_FY1_YEAR, 3,
                          f"=YEAR(IQ_FY1_FYE_DATE({TICKER_REF}))"))
    style_formula(ws.cell(broker_layout.ROW_FY2_YEAR, 3,
                          f"=YEAR(IQ_FY2_FYE_DATE({TICKER_REF}))"))
    style_formula(ws.cell(broker_layout.ROW_FY3_YEAR, 3,
                          f"=YEAR(IQ_FY3_FYE_DATE({TICKER_REF}))"))

    for j, h in enumerate(broker_layout.COL_HEADERS):
        style_header(ws.cell(broker_layout.ROW_COL_HEADERS, j + 2, h))

    # Section A: P&L estimates. Cols C/D/E means, F/G high/low, H count.
    for row, label, mean_fn, high_fn, low_fn, count_fn, omit_ccy in broker_layout.PNL:
        ws.cell(row, 2, label).font = LABEL
        for col_letter, period in zip(broker_layout.MEAN_COLS, broker_layout.PERIODS):
            cell = ws[f"{col_letter}{row}"]
            cell.value = f"={mean_fn}({_est_args(TICKER_REF, period, omit_ccy)})"
            style_formula(cell)
        cell = ws[f"{broker_layout.HIGH_COL}{row}"]
        cell.value = f"={high_fn}({_est_args(TICKER_REF, 'IQ_FY1', omit_ccy)})"
        style_formula(cell)
        cell = ws[f"{broker_layout.LOW_COL}{row}"]
        cell.value = f"={low_fn}({_est_args(TICKER_REF, 'IQ_FY1', omit_ccy)})"
        style_formula(cell)
        cell = ws[f"{broker_layout.COUNT_COL}{row}"]
        cell.value = f"={count_fn}({TICKER_REF},IQ_FY1)"
        style_formula(cell)

    # Section B: implied calcs — left blank in fetcher; main template has the formulas.
    for r, label, _ in broker_layout.IMPLIED:
        ws.cell(r, 2, label).font = LABEL

    # Section C: sentiment
    for r, label, formula_template in broker_layout.SENTIMENT:
        ws.cell(r, 2, label).font = LABEL
        if formula_template is None:
            continue
        formula = formula_template.format(t=TICKER_REF)
        style_formula(ws.cell(r, 3, formula))

    ws.column_dimensions["B"].width = 38
    for letter in ["C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[letter].width = 14


def build():
    wb = Workbook()
    wb.remove(wb.active)
    fetcher = wb.create_sheet("Fetcher")
    build_fetcher(fetcher)

    # Ticker now lives in column C (column B holds the label).
    wb.defined_names[TICKER_REF] = DefinedName(
        name=TICKER_REF,
        attr_text=f"Fetcher!$C${broker_layout.ROW_TICKER}",
    )

    FETCHER_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(FETCHER_PATH)
    return FETCHER_PATH


if __name__ == "__main__":
    out = build()
    print(f"Wrote {out}")
