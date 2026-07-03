"""Scaffold templates/multiple_history_fetcher.xlsx.

Idempotent. Creates a single-tab workbook holding live CapIQ formulas for
~6 years of daily NTM EV/EBITDA history. Driven at runtime by
companies/scripts/fetch_multiple_history.py.

Run:
    python -m shared.scaffold_multiple_history_fetcher
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string
from openpyxl.workbook.defined_name import DefinedName

from shared import multiple_history_layout as layout
# Styling constants + style_* helpers are shared across all scaffolders.
from shared.excel_helpers import (
    ARIAL, ARIAL_SIZE, BANNER_FONT, FORMULA_FONT, LABEL, LABEL_BOLD,
    TITLE_FONT, style_header, style_input,
)

REPO_ROOT = Path(__file__).resolve().parent.parent
FETCHER_PATH = REPO_ROOT / "templates" / "multiple_history_fetcher.xlsx"

GRAY_CALC = "666666"
CALC_FONT = Font(color=GRAY_CALC, name=ARIAL, size=ARIAL_SIZE)


def build_fetcher(ws):
    ws.sheet_view.showGridLines = False

    # Universal: column A spacer + B2 dynamic title.
    for letter, width in layout.COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    title = ws.cell(layout.ROW_TITLE, 2,
                    f'={layout.NAME_TICKER} & " | NTM EV/EBITDA History"')
    title.font = TITLE_FONT

    ws.cell(layout.ROW_BANNER, 2,
            "Historical NTM EV/EBITDA Multiple — formulas only. "
            "Open in Excel with CapIQ plugin loaded to refresh.").font = BANNER_FONT

    ws.cell(layout.ROW_RUN_VIA, 2, "Run via:").font = LABEL_BOLD
    ws.cell(layout.ROW_RUN_VIA, 3,
            "python -m companies.scripts.fetch_multiple_history <TICKER>").font = LABEL

    # Inputs: ticker, end date, lookback years.
    ws.cell(layout.ROW_TICKER, 2, "Ticker:").font = LABEL_BOLD
    ticker_cell = ws.cell(layout.ROW_TICKER, 3, "AAPL")
    style_input(ticker_cell)
    ticker_cell.comment = Comment(
        "Set the ticker you want history for. fetch_multiple_history.py drives "
        "this programmatically; you can also change it manually for ad-hoc refreshes.",
        "scaffold_multiple_history_fetcher",
    )

    ws.cell(layout.ROW_END_DATE, 2, "End Date:").font = LABEL_BOLD
    end_cell = ws.cell(layout.ROW_END_DATE, 3, "=TODAY()")
    style_input(end_cell)
    end_cell.number_format = "mm/dd/yyyy"

    ws.cell(layout.ROW_LOOKBACK_YRS, 2, "Lookback Years:").font = LABEL_BOLD
    lb_cell = ws.cell(layout.ROW_LOOKBACK_YRS, 3, 5)
    style_input(lb_cell)
    lb_cell.number_format = "0"

    # Column headers (ROW_COL_HEADERS, columns B..{layout.DATA_COLUMNS[-1][0]}).
    for letter, header, _role in layout.DATA_COLUMNS:
        col_idx = column_index_from_string(letter)
        style_header(ws.cell(layout.ROW_COL_HEADERS, col_idx, header))

    # Data rows: input formulas in CapIQ columns; calc formulas in derived
    # columns. Date column left blank — populated at fetch time.
    for r in range(layout.ROW_DATA_START, layout.ROW_DATA_END + 1):
        for letter, _label, role in layout.DATA_COLUMNS:
            col_idx = column_index_from_string(letter)
            cell = ws.cell(r, col_idx)
            cell.number_format = layout.COLUMN_NUMBER_FORMATS[letter]
            if role == "date":
                cell.font = LABEL  # plain Arial 10
            elif role == "input_formula":
                template = layout.INPUT_FORMULAS[letter]
                cell.value = template.format(t=layout.NAME_TICKER, d=f"B{r}")
                cell.font = FORMULA_FONT
            elif role == "calc":
                template = layout.CALC_FORMULAS[letter]
                cell.value = template.format(r=r)
                cell.font = CALC_FONT

    ws.freeze_panes = "A11"


def build():
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Fetcher")
    build_fetcher(ws)

    wb.defined_names[layout.NAME_TICKER] = DefinedName(
        name=layout.NAME_TICKER,
        attr_text=f"Fetcher!$C${layout.ROW_TICKER}",
    )
    wb.defined_names[layout.NAME_END_DATE] = DefinedName(
        name=layout.NAME_END_DATE,
        attr_text=f"Fetcher!$C${layout.ROW_END_DATE}",
    )
    wb.defined_names[layout.NAME_LOOKBACK_YRS] = DefinedName(
        name=layout.NAME_LOOKBACK_YRS,
        attr_text=f"Fetcher!$C${layout.ROW_LOOKBACK_YRS}",
    )

    FETCHER_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(FETCHER_PATH)
    return FETCHER_PATH


if __name__ == "__main__":
    out = build()
    print(f"Wrote {out}")
