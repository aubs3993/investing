"""Scaffold templates/company_model.xlsx — base corporate financial model.

Idempotent: re-running overwrites the file. This is the master template only —
runtime scripts copy this file to companies/output/<TICKER>/ before filling.

Universal layout conventions (apply to every tab):
  - Gridlines off
  - Default font Arial 10pt (applied via cell styling)
  - Column A is a 2.71-wide blank spacer; all content starts in column B
  - Row 1 is blank; row 2 holds a dynamic title formula at B2
    (=inp_ticker & " | <Tab Title>"); row 3 is blank; content begins at row 4

Run from repo root:
    python -m shared.scaffold_template
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from shared import broker_layout, capiq_layout


REPO_ROOT = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = REPO_ROOT / "templates" / "company_model.xlsx"


# --- Universal styling (Arial 10pt across the workbook) ---
ARIAL = "Arial"
ARIAL_SIZE = 10

BLUE_INPUT = "0000FF"
GREEN_LINK = "008000"
BLACK = "000000"
WHITE = "FFFFFF"
YELLOW_FILL = "FFFF99"
HEADER_BLUE = "0070C0"
BANNER_GRAY = "808080"

HAIR = Side(border_style="hair")
THIN_TOP = Side(border_style="thin")

INPUT_FONT = Font(name=ARIAL, size=ARIAL_SIZE, bold=False, color=BLUE_INPUT)
INPUT_FILL = PatternFill("solid", fgColor=YELLOW_FILL)
INPUT_BORDER = Border(top=HAIR, bottom=HAIR, left=HAIR, right=HAIR)

HEADER_FONT = Font(name=ARIAL, size=ARIAL_SIZE, bold=True, color=WHITE)
HEADER_FILL = PatternFill("solid", fgColor=HEADER_BLUE)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

LINK_FONT = Font(name=ARIAL, size=ARIAL_SIZE, color=GREEN_LINK)
FORMULA_FONT = Font(name=ARIAL, size=ARIAL_SIZE, color=BLACK)
SUBTOTAL_FONT = Font(name=ARIAL, size=ARIAL_SIZE, bold=True, color=BLACK)
SUBTOTAL_TOP_BORDER = Border(top=THIN_TOP)

TITLE_FONT = Font(name=ARIAL, size=ARIAL_SIZE, bold=True, color=BLACK)
LABEL_FONT = Font(name=ARIAL, size=ARIAL_SIZE, color=BLACK)
LABEL_BOLD_FONT = Font(name=ARIAL, size=ARIAL_SIZE, bold=True, color=BLACK)
SECTION_FONT = Font(name=ARIAL, size=ARIAL_SIZE, bold=True, color=BLACK)
BANNER_FONT = Font(name=ARIAL, size=ARIAL_SIZE, italic=True, color=BANNER_GRAY)


# --- Number formats ---
NUM = "#,##0;(#,##0)"
NUM_DEC = "#,##0.00;(#,##0.00)"
PCT = "0.0%"
MULT = '0.0"x"'
RATIO = '0.00"x"'
DATE = "mm/dd/yyyy"
DATETIME = "mm/dd/yyyy hh:mm"
PRICE = "$#,##0.00;($#,##0.00)"
DPS = "$0.00"
TEXT = "@"


# --- Style helpers ---
def style_input(cell, num_format=NUM):
    cell.font = INPUT_FONT
    cell.fill = INPUT_FILL
    cell.border = INPUT_BORDER
    cell.number_format = num_format


def style_formula(cell, num_format=NUM):
    cell.font = FORMULA_FONT
    cell.number_format = num_format


def style_link(cell, num_format=NUM):
    cell.font = LINK_FONT
    cell.number_format = num_format


def style_subtotal(cell, num_format=NUM):
    cell.font = SUBTOTAL_FONT
    cell.border = SUBTOTAL_TOP_BORDER
    cell.number_format = num_format


def style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN


def style_label(cell, bold=False):
    cell.font = LABEL_BOLD_FONT if bold else LABEL_FONT


def add_named_range(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name=name, attr_text=f"{sheet}!{ref}")


# --- Universal sheet setup ---
def apply_sheet_defaults(ws, tab_title: str):
    """Apply gridlines-off, column A spacer, blank row 1, and B2 title."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2.71
    title_cell = ws.cell(2, 2, f'=inp_ticker & " | {tab_title}"')
    title_cell.font = TITLE_FONT


def make_year_columns():
    cy = datetime.now().year
    return (
        [(cy + o, f"{cy + o}A") for o in (-3, -2, -1)]
        + [(cy + o, f"{cy + o}E") for o in range(0, 10)]
    )


def col(i):
    """0-indexed year position -> Excel column letter (C..O).

    Year columns sit to the right of column A (spacer) and column B (label),
    so the first year column is C (index 0 -> C).
    """
    return get_column_letter(i + 3)


# --- Inputs tab layout (referenced by other tabs via named ranges) ---
# Universal: row 1 blank, row 2 title, row 3 blank, content from row 4.
INPUTS_SECTION_A_ROW = 4
INPUTS_FIRST_ROW = 5

# Tuple: (label, named_range, num_format, default, link_formula)
# link_formula present -> green LINK; otherwise blue INPUT cell.
INPUT_ROWS = [
    ("Ticker",                     "inp_ticker",              TEXT,  "TICKR",        None),
    ("Run Date",                   "inp_run_date",            DATE,  None,           None),
    ("Fiscal Year End Month",      "inp_fye_month",           "0",   12,             None),
    ("Company Name",               "inp_company_name",        TEXT,  None,           "=_CapIQ_Data!F12"),
    ("Sector",                     "inp_sector",              TEXT,  None,           "=_CapIQ_Data!F13"),
    ("Currency",                   "inp_currency",            TEXT,  None,           "=_CapIQ_Data!F14"),
    ("Current Price",              "inp_current_price",       PRICE, None,           "=_CapIQ_Data!F18"),
    ("Diluted Shares Outstanding", "inp_diluted_shares",      NUM,   None,           "=_CapIQ_Data!F19"),
    ("Annual DPS",                 "inp_annual_dps",          DPS,   None,           "=IFERROR(_CapIQ_Data!E40, _CapIQ_Data!D40)"),
    ("DPS Annual Growth %",        "inp_dps_growth",          PCT,   0,              None),
    ("Tax Rate",                   "inp_tax_rate",            PCT,   0.21,           None),
    ("Cash & Equivalents",         "inp_cash",                NUM,   None,           "=_CapIQ_Data!F21"),
    ("Total Debt",                 "inp_debt",                NUM,   None,           "=_CapIQ_Data!F23"),
    ("Minority Interest",          "inp_minority_interest",   NUM,   None,           "=_CapIQ_Data!F25"),
    ("Equity Investments",         "inp_equity_investments",  NUM,   None,           "=_CapIQ_Data!F26"),
    ("Cash Sweep %",               "inp_cash_sweep_pct",      PCT,   0,              None),
    ("Minimum Cash Balance",       "inp_min_cash",            NUM,   0,              None),
]

INPUTS_SECTION_B_ROW = INPUTS_FIRST_ROW + len(INPUT_ROWS) + 1  # 5 + 17 + 1 = 23
CALC_ROWS = [
    ("Annual Dividends Paid (current)", "=inp_annual_dps*inp_diluted_shares",                                            "calc_annual_div", NUM),
    ("Current Market Cap",              "=inp_current_price*inp_diluted_shares",                                         "mkt_cap",         NUM),
    ("Net Debt",                        "=inp_debt-inp_cash",                                                            None,              NUM),
    ("Current Enterprise Value",        "=mkt_cap+inp_debt-inp_cash+inp_minority_interest-inp_equity_investments",       None,              NUM),
]

INPUTS_SECTION_C_ROW = INPUTS_SECTION_B_ROW + len(CALC_ROWS) + 2  # 23 + 4 + 2 = 29
DRIVER_HEADER_ROW = INPUTS_SECTION_C_ROW + 1  # 30
DRIVER_FIRST_ROW = INPUTS_SECTION_C_ROW + 2   # 31
DRIVERS = [
    ("Revenue Growth %",         PCT,  0.05),
    ("Gross Margin %",           PCT,  0.40),
    ("Total OpEx % of Revenue",  PCT,  0.20),
    ("CapEx % of Revenue",       PCT,  0.05),
    ("D&A % of CapEx",           RATIO, 1.0),
    ("Exit EBITDA Multiple",     MULT, 10),
]
DRV_REV   = DRIVER_FIRST_ROW + 0
DRV_GM    = DRIVER_FIRST_ROW + 1
DRV_OPEX  = DRIVER_FIRST_ROW + 2
DRV_CAPEX = DRIVER_FIRST_ROW + 3
DRV_DA    = DRIVER_FIRST_ROW + 4
DRV_EXIT  = DRIVER_FIRST_ROW + 5


# --- IS row map (years are columns C..O = 13 cols starting at col index 3) ---
IS_ROW_HEADER  = 4
IS_ROW_REV     = 5
IS_ROW_GROWTH  = 6
IS_ROW_COGS    = 7
IS_ROW_GP      = 8
IS_ROW_GM      = 9
IS_ROW_OPEX    = 10
IS_ROW_EBITDA  = 11
IS_ROW_EBITDAM = 12
IS_ROW_DA      = 13
IS_ROW_EBIT    = 14
IS_ROW_EBITM   = 15
IS_ROW_INTEXP  = 16
IS_ROW_INTINC  = 17
IS_ROW_PRETAX  = 18
IS_ROW_TAX     = 19
IS_ROW_ETR     = 20
IS_ROW_NI      = 21
IS_ROW_SHARES  = 22
IS_ROW_EPS     = 23

# --- CF row map ---
CF_ROW_HEADER     = 4
CF_ROW_EBITDA     = 5
CF_ROW_TAXES      = 6
CF_ROW_INTEREST   = 7
CF_ROW_CAPEX      = 8
CF_ROW_LFCF       = 9
CF_ROW_DEBT_AMORT = 10
CF_ROW_DIVIDENDS  = 11
CF_ROW_NETCHG     = 12
CF_ROW_BEGCASH    = 13
CF_ROW_ENDCASH    = 14

# --- Debt row map ---
DEBT_ROW_HEADER    = 4
DEBT_ROW_REVOLVER  = 5
DEBT_ROW_REV_BEG   = 6
DEBT_ROW_REV_DRAW  = 7
DEBT_ROW_REV_REPAY = 8
DEBT_ROW_REV_END   = 9
DEBT_ROW_REV_AVG   = 10
DEBT_ROW_REV_RATE  = 11
DEBT_ROW_REV_INT   = 12
DEBT_ROW_TL_TITLE  = 14
DEBT_ROW_TL_BEG    = 15
DEBT_ROW_TL_AMORT  = 16
DEBT_ROW_TL_PREPAY = 17
DEBT_ROW_TL_END    = 18
DEBT_ROW_TL_AVG    = 19
DEBT_ROW_TL_RATE   = 20
DEBT_ROW_TL_INT    = 21
DEBT_ROW_SN_TITLE  = 23
DEBT_ROW_SN_BEG    = 24
DEBT_ROW_SN_REPAY  = 25
DEBT_ROW_SN_END    = 26
DEBT_ROW_SN_RATE   = 27
DEBT_ROW_SN_INT    = 28
DEBT_ROW_TOT_TITLE = 30
DEBT_ROW_TOT_BEG   = 31
DEBT_ROW_TOT_END   = 32
DEBT_ROW_TOT_INT   = 33
DEBT_ROW_TOT_AMORT = 34
DEBT_ROW_TOT_PREP  = 35
DEBT_ROW_SW_TITLE  = 37
DEBT_ROW_SW_AVAIL  = 38
DEBT_ROW_SW_APPLY  = 39
DEBT_ROW_SW_TL     = 40

# --- Valuation row map ---
VAL_ROW_HEADER  = 4
VAL_ROW_EBITDA  = 5
VAL_ROW_MULT    = 6
VAL_ROW_EV      = 7
VAL_ROW_DEBT    = 8
VAL_ROW_CASH    = 9
VAL_ROW_NCI     = 10
VAL_ROW_EQINV   = 11
VAL_ROW_EQUITY  = 12
VAL_ROW_SHARES  = 13
VAL_ROW_PPS     = 14
VAL_ROW_IRR     = 15
VAL_ROW_MOIC    = 16


# --- Build functions ---
def build_inputs(ws, years):
    apply_sheet_defaults(ws, "Inputs & Assumptions")

    style_label(ws.cell(INPUTS_SECTION_A_ROW, 2,
                        "Inputs (CapIQ-driven where available; manual otherwise)"), bold=True)
    for i, (label, _name, fmt, default, link) in enumerate(INPUT_ROWS):
        r = INPUTS_FIRST_ROW + i
        style_label(ws.cell(r, 2, label))
        if link is not None:
            c = ws.cell(r, 3, link)
            style_link(c, num_format=fmt)
        else:
            val = default if default is not None else (datetime.now().date() if fmt == DATE else 0)
            c = ws.cell(r, 3, val)
            style_input(c, num_format=fmt)

    style_label(ws.cell(INPUTS_SECTION_B_ROW, 2, "Calculated"), bold=True)
    for i, (label, formula, _name, fmt) in enumerate(CALC_ROWS):
        r = INPUTS_SECTION_B_ROW + 1 + i
        style_label(ws.cell(r, 2, label))
        c = ws.cell(r, 3, formula)
        style_formula(c, num_format=fmt)

    style_label(ws.cell(INPUTS_SECTION_C_ROW, 2, "Driver Assumptions (manual fill)"), bold=True)
    for i, (_yr, lbl) in enumerate(years):
        if lbl.endswith("E"):
            style_header(ws.cell(DRIVER_HEADER_ROW, i + 3, lbl))
    for i, (label, fmt, default) in enumerate(DRIVERS):
        r = DRIVER_FIRST_ROW + i
        style_label(ws.cell(r, 2, label))
        for j, (_yr, lbl) in enumerate(years):
            if lbl.endswith("E"):
                c = ws.cell(r, j + 3, default)
                style_input(c, num_format=fmt)

    ws.column_dimensions["B"].width = 35
    for j in range(len(years)):
        ws.column_dimensions[col(j)].width = 14


def register_inputs_named_ranges(wb):
    """Inputs values live in column C (was column B before the universal shift)."""
    for i, (_label, name, _fmt, _default, _link) in enumerate(INPUT_ROWS):
        r = INPUTS_FIRST_ROW + i
        add_named_range(wb, name, "Inputs", f"$C${r}")
    for i, (_label, _formula, name, _fmt) in enumerate(CALC_ROWS):
        if name is None:
            continue
        r = INPUTS_SECTION_B_ROW + 1 + i
        add_named_range(wb, name, "Inputs", f"$C${r}")


def register_driver_named_ranges(wb):
    """Driver assumption rows on Inputs (F..O for the 10 projection years)."""
    drv_ranges = {
        "drv_revenue_growth": DRV_REV,
        "drv_gross_margin":   DRV_GM,
        "drv_opex_pct_rev":   DRV_OPEX,
        "drv_capex_pct_rev":  DRV_CAPEX,
        "drv_da_pct_capex":   DRV_DA,
        "drv_exit_multiple":  DRV_EXIT,
    }
    for name, row in drv_ranges.items():
        add_named_range(wb, name, "Inputs", f"$F${row}:$O${row}")


def build_cover(ws, _years):
    apply_sheet_defaults(ws, "Cover")
    rows = [
        ("Company Name",  "=inp_company_name",  TEXT),
        ("Ticker",        "=inp_ticker",        TEXT),
        ("Sector",        "=inp_sector",        TEXT),
        ("Currency",      "=inp_currency",      TEXT),
        ("Run Date",      "=inp_run_date",      DATE),
        ("Current Price", "=inp_current_price", PRICE),
        ("Market Cap",    "=mkt_cap",           NUM),
    ]
    for i, (label, formula, fmt) in enumerate(rows):
        r = 4 + i
        style_label(ws.cell(r, 2, label), bold=True)
        c = ws.cell(r, 3, formula)
        style_link(c, num_format=fmt)
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28


def _is_proj(years, i):
    return years[i][1].endswith("E")


def _set_year_headers(ws, years, header_row, projection_only=False):
    for i, (_yr, lbl) in enumerate(years):
        if projection_only and not lbl.endswith("E"):
            continue
        style_header(ws.cell(header_row, i + 3, lbl))


def _set_widths(ws, n):
    ws.column_dimensions["B"].width = 35
    for i in range(n):
        ws.column_dimensions[col(i)].width = 14


def build_is(ws, years):
    apply_sheet_defaults(ws, "Income Statement")
    n = len(years)
    _set_year_headers(ws, years, IS_ROW_HEADER)

    style_label(ws.cell(IS_ROW_REV, 2, "Revenue"))
    for i in range(n):
        cur = col(i)
        if not _is_proj(years, i):
            style_link(ws.cell(IS_ROW_REV, i + 3, f"=_CapIQ_Data!{cur}31"))
        else:
            prev = col(i - 1)
            style_formula(ws.cell(IS_ROW_REV, i + 3, f"={prev}{IS_ROW_REV}*(1+Inputs!{cur}${DRV_REV})"))

    style_label(ws.cell(IS_ROW_GROWTH, 2, "Revenue Growth %"))
    for i in range(n):
        cur = col(i)
        if i == 0:
            ws.cell(IS_ROW_GROWTH, i + 3, None)
        else:
            prev = col(i - 1)
            style_formula(
                ws.cell(IS_ROW_GROWTH, i + 3, f"={cur}{IS_ROW_REV}/{prev}{IS_ROW_REV}-1"),
                num_format=PCT,
            )

    style_label(ws.cell(IS_ROW_COGS, 2, "COGS"))
    for i in range(n):
        cur = col(i)
        if not _is_proj(years, i):
            style_link(ws.cell(IS_ROW_COGS, i + 3, f"=_CapIQ_Data!{cur}32"))
        else:
            style_formula(ws.cell(IS_ROW_COGS, i + 3, f"={cur}{IS_ROW_REV}*(1-Inputs!{cur}${DRV_GM})"))

    style_label(ws.cell(IS_ROW_GP, 2, "Gross Profit"))
    for i in range(n):
        cur = col(i)
        style_subtotal(ws.cell(IS_ROW_GP, i + 3, f"={cur}{IS_ROW_REV}-{cur}{IS_ROW_COGS}"))

    style_label(ws.cell(IS_ROW_GM, 2, "Gross Margin %"))
    for i in range(n):
        cur = col(i)
        style_formula(
            ws.cell(IS_ROW_GM, i + 3, f"=IFERROR({cur}{IS_ROW_GP}/{cur}{IS_ROW_REV},0)"),
            num_format=PCT,
        )

    style_label(ws.cell(IS_ROW_OPEX, 2, "Total OpEx"))
    for i in range(n):
        cur = col(i)
        if not _is_proj(years, i):
            style_link(ws.cell(IS_ROW_OPEX, i + 3, f"=_CapIQ_Data!{cur}34"))
        else:
            style_formula(ws.cell(IS_ROW_OPEX, i + 3, f"={cur}{IS_ROW_REV}*Inputs!{cur}${DRV_OPEX}"))

    style_label(ws.cell(IS_ROW_EBITDA, 2, "EBITDA"))
    for i in range(n):
        cur = col(i)
        style_subtotal(ws.cell(IS_ROW_EBITDA, i + 3, f"={cur}{IS_ROW_GP}-{cur}{IS_ROW_OPEX}"))

    style_label(ws.cell(IS_ROW_EBITDAM, 2, "EBITDA Margin %"))
    for i in range(n):
        cur = col(i)
        style_formula(
            ws.cell(IS_ROW_EBITDAM, i + 3, f"=IFERROR({cur}{IS_ROW_EBITDA}/{cur}{IS_ROW_REV},0)"),
            num_format=PCT,
        )

    style_label(ws.cell(IS_ROW_DA, 2, "D&A"))
    for i in range(n):
        cur = col(i)
        if not _is_proj(years, i):
            style_link(ws.cell(IS_ROW_DA, i + 3, f"=_CapIQ_Data!{cur}35"))
        else:
            # CapEx on CF is negative; negate, then multiply by D&A%
            style_link(
                ws.cell(IS_ROW_DA, i + 3, f"=-CF!{cur}{CF_ROW_CAPEX}*Inputs!{cur}${DRV_DA}")
            )

    style_label(ws.cell(IS_ROW_EBIT, 2, "EBIT"))
    for i in range(n):
        cur = col(i)
        style_subtotal(ws.cell(IS_ROW_EBIT, i + 3, f"={cur}{IS_ROW_EBITDA}-{cur}{IS_ROW_DA}"))

    style_label(ws.cell(IS_ROW_EBITM, 2, "EBIT Margin %"))
    for i in range(n):
        cur = col(i)
        style_formula(
            ws.cell(IS_ROW_EBITM, i + 3, f"=IFERROR({cur}{IS_ROW_EBIT}/{cur}{IS_ROW_REV},0)"),
            num_format=PCT,
        )

    # Per spec: leave historical cells blank for IntExp / IntInc / Pretax / Tax / NI.
    style_label(ws.cell(IS_ROW_INTEXP, 2, "Interest Expense"))
    for i in range(n):
        cur = col(i)
        if not _is_proj(years, i):
            ws.cell(IS_ROW_INTEXP, i + 3, None)
        else:
            style_link(ws.cell(IS_ROW_INTEXP, i + 3, f"=Debt!{cur}{DEBT_ROW_TOT_INT}"))

    style_label(ws.cell(IS_ROW_INTINC, 2, "Interest Income"))
    first_proj = next(j for j, (_y, l) in enumerate(years) if l.endswith("E"))
    for i in range(n):
        cur = col(i)
        if not _is_proj(years, i):
            ws.cell(IS_ROW_INTINC, i + 3, None)
        elif i == first_proj:
            style_input(ws.cell(IS_ROW_INTINC, i + 3, 0))
        else:
            prev = col(i - 1)
            style_formula(ws.cell(IS_ROW_INTINC, i + 3, f"={prev}{IS_ROW_INTINC}"))

    style_label(ws.cell(IS_ROW_PRETAX, 2, "Pre-tax Income"))
    for i in range(n):
        cur = col(i)
        if not _is_proj(years, i):
            ws.cell(IS_ROW_PRETAX, i + 3, None)
        else:
            style_subtotal(
                ws.cell(IS_ROW_PRETAX, i + 3,
                        f"={cur}{IS_ROW_EBIT}-{cur}{IS_ROW_INTEXP}+{cur}{IS_ROW_INTINC}")
            )

    style_label(ws.cell(IS_ROW_TAX, 2, "Taxes"))
    for i in range(n):
        cur = col(i)
        if not _is_proj(years, i):
            ws.cell(IS_ROW_TAX, i + 3, None)
        else:
            style_formula(ws.cell(IS_ROW_TAX, i + 3, f"={cur}{IS_ROW_PRETAX}*inp_tax_rate"))

    style_label(ws.cell(IS_ROW_ETR, 2, "Effective Tax Rate %"))
    for i in range(n):
        cur = col(i)
        if not _is_proj(years, i):
            ws.cell(IS_ROW_ETR, i + 3, None)
        else:
            style_formula(
                ws.cell(IS_ROW_ETR, i + 3,
                        f"=IFERROR({cur}{IS_ROW_TAX}/{cur}{IS_ROW_PRETAX},0)"),
                num_format=PCT,
            )

    style_label(ws.cell(IS_ROW_NI, 2, "Net Income"))
    for i in range(n):
        cur = col(i)
        if not _is_proj(years, i):
            ws.cell(IS_ROW_NI, i + 3, None)
        else:
            style_subtotal(
                ws.cell(IS_ROW_NI, i + 3, f"={cur}{IS_ROW_PRETAX}-{cur}{IS_ROW_TAX}")
            )

    # Diluted shares: hold flat across all 13 columns (no historical link).
    style_label(ws.cell(IS_ROW_SHARES, 2, "Diluted Shares Outstanding"))
    for i in range(n):
        style_link(ws.cell(IS_ROW_SHARES, i + 3, "=inp_diluted_shares"))

    style_label(ws.cell(IS_ROW_EPS, 2, "Diluted EPS"))
    for i in range(n):
        cur = col(i)
        style_formula(
            ws.cell(IS_ROW_EPS, i + 3, f"=IFERROR({cur}{IS_ROW_NI}/{cur}{IS_ROW_SHARES},0)"),
            num_format=NUM_DEC,
        )

    ws.freeze_panes = "C5"
    _set_widths(ws, n)


def build_cf(ws, years):
    apply_sheet_defaults(ws, "Cash Flow")
    n = len(years)
    _set_year_headers(ws, years, CF_ROW_HEADER)

    style_label(ws.cell(CF_ROW_EBITDA, 2, "EBITDA"))
    for i in range(n):
        cur = col(i)
        style_link(ws.cell(CF_ROW_EBITDA, i + 3, f"=IS!{cur}{IS_ROW_EBITDA}"))

    style_label(ws.cell(CF_ROW_TAXES, 2, "(Less) Cash Taxes"))
    for i in range(n):
        cur = col(i)
        if not _is_proj(years, i):
            ws.cell(CF_ROW_TAXES, i + 3, None)
        else:
            style_link(ws.cell(CF_ROW_TAXES, i + 3, f"=-IS!{cur}{IS_ROW_TAX}"))

    style_label(ws.cell(CF_ROW_INTEREST, 2, "(Less) Cash Interest"))
    for i in range(n):
        cur = col(i)
        if not _is_proj(years, i):
            ws.cell(CF_ROW_INTEREST, i + 3, None)
        else:
            style_link(ws.cell(CF_ROW_INTEREST, i + 3, f"=-IS!{cur}{IS_ROW_INTEXP}"))

    style_label(ws.cell(CF_ROW_CAPEX, 2, "(Less) CapEx"))
    for i in range(n):
        cur = col(i)
        if not _is_proj(years, i):
            style_link(ws.cell(CF_ROW_CAPEX, i + 3, f"=-_CapIQ_Data!{cur}38"))
        else:
            style_formula(
                ws.cell(CF_ROW_CAPEX, i + 3, f"=-IS!{cur}{IS_ROW_REV}*Inputs!{cur}${DRV_CAPEX}")
            )

    style_label(ws.cell(CF_ROW_LFCF, 2, "Levered Free Cash Flow"))
    for i in range(n):
        cur = col(i)
        style_subtotal(
            ws.cell(CF_ROW_LFCF, i + 3,
                    f"=SUM({cur}{CF_ROW_EBITDA}:{cur}{CF_ROW_CAPEX})")
        )

    style_label(ws.cell(CF_ROW_DEBT_AMORT, 2, "(Less) Debt Amortization"))
    for i in range(n):
        cur = col(i)
        if not _is_proj(years, i):
            ws.cell(CF_ROW_DEBT_AMORT, i + 3, None)
        else:
            style_link(
                ws.cell(CF_ROW_DEBT_AMORT, i + 3,
                        f"=-(Debt!{cur}{DEBT_ROW_TOT_AMORT}+Debt!{cur}{DEBT_ROW_TOT_PREP})")
            )

    style_label(ws.cell(CF_ROW_DIVIDENDS, 2, "(Less) Dividends"))
    first_proj = next(j for j, (_y, l) in enumerate(years) if l.endswith("E"))
    for i in range(n):
        cur = col(i)
        if not _is_proj(years, i):
            style_input(ws.cell(CF_ROW_DIVIDENDS, i + 3, 0))
        else:
            k = i - first_proj  # 0 for first projection year
            f = f"=-inp_annual_dps*((1+inp_dps_growth)^{k})*IS!{cur}{IS_ROW_SHARES}"
            style_formula(ws.cell(CF_ROW_DIVIDENDS, i + 3, f))

    style_label(ws.cell(CF_ROW_NETCHG, 2, "Net Change in Cash"))
    for i in range(n):
        cur = col(i)
        style_subtotal(
            ws.cell(CF_ROW_NETCHG, i + 3,
                    f"={cur}{CF_ROW_LFCF}+{cur}{CF_ROW_DEBT_AMORT}+{cur}{CF_ROW_DIVIDENDS}")
        )

    style_label(ws.cell(CF_ROW_BEGCASH, 2, "Beginning Cash"))
    for i in range(n):
        cur = col(i)
        if i == 0:
            style_input(ws.cell(CF_ROW_BEGCASH, i + 3, 0))
        else:
            prev = col(i - 1)
            style_formula(ws.cell(CF_ROW_BEGCASH, i + 3, f"={prev}{CF_ROW_ENDCASH}"))

    style_label(ws.cell(CF_ROW_ENDCASH, 2, "Ending Cash"))
    for i in range(n):
        cur = col(i)
        style_subtotal(
            ws.cell(CF_ROW_ENDCASH, i + 3,
                    f"={cur}{CF_ROW_BEGCASH}+{cur}{CF_ROW_NETCHG}")
        )

    ws.freeze_panes = "C5"
    _set_widths(ws, n)


def build_debt(ws, years):
    apply_sheet_defaults(ws, "Debt Schedule")
    n = len(years)
    _set_year_headers(ws, years, DEBT_ROW_HEADER)

    style_label(ws.cell(DEBT_ROW_REVOLVER, 2, "Revolver"), bold=True)
    style_label(ws.cell(DEBT_ROW_REV_BEG,   2, "Beginning Balance"))
    style_label(ws.cell(DEBT_ROW_REV_DRAW,  2, "Draws"))
    style_label(ws.cell(DEBT_ROW_REV_REPAY, 2, "Repayments"))
    style_label(ws.cell(DEBT_ROW_REV_END,   2, "Ending Balance"))
    style_label(ws.cell(DEBT_ROW_REV_AVG,   2, "Average Balance"))
    style_label(ws.cell(DEBT_ROW_REV_RATE,  2, "Interest Rate %"))
    style_label(ws.cell(DEBT_ROW_REV_INT,   2, "Interest Expense"))
    for i in range(n):
        cur = col(i)
        if i == 0:
            style_input(ws.cell(DEBT_ROW_REV_BEG, i + 3, 0))
        else:
            prev = col(i - 1)
            style_formula(ws.cell(DEBT_ROW_REV_BEG, i + 3, f"={prev}{DEBT_ROW_REV_END}"))
        style_input(ws.cell(DEBT_ROW_REV_DRAW, i + 3, 0))
        style_input(ws.cell(DEBT_ROW_REV_REPAY, i + 3, 0))
        style_subtotal(
            ws.cell(DEBT_ROW_REV_END, i + 3,
                    f"={cur}{DEBT_ROW_REV_BEG}+{cur}{DEBT_ROW_REV_DRAW}-{cur}{DEBT_ROW_REV_REPAY}")
        )
        style_formula(
            ws.cell(DEBT_ROW_REV_AVG, i + 3,
                    f"=({cur}{DEBT_ROW_REV_BEG}+{cur}{DEBT_ROW_REV_END})/2")
        )
        style_input(ws.cell(DEBT_ROW_REV_RATE, i + 3, 0.06), num_format=PCT)
        style_formula(
            ws.cell(DEBT_ROW_REV_INT, i + 3,
                    f"={cur}{DEBT_ROW_REV_AVG}*{cur}{DEBT_ROW_REV_RATE}")
        )

    style_label(ws.cell(DEBT_ROW_TL_TITLE, 2, "Term Loan"), bold=True)
    style_label(ws.cell(DEBT_ROW_TL_BEG,    2, "Beginning Balance"))
    style_label(ws.cell(DEBT_ROW_TL_AMORT,  2, "Mandatory Amortization"))
    style_label(ws.cell(DEBT_ROW_TL_PREPAY, 2, "Optional Prepayment (Cash Sweep)"))
    style_label(ws.cell(DEBT_ROW_TL_END,    2, "Ending Balance"))
    style_label(ws.cell(DEBT_ROW_TL_AVG,    2, "Average Balance"))
    style_label(ws.cell(DEBT_ROW_TL_RATE,   2, "Interest Rate %"))
    style_label(ws.cell(DEBT_ROW_TL_INT,    2, "Interest Expense"))
    for i in range(n):
        cur = col(i)
        if i == 0:
            style_input(ws.cell(DEBT_ROW_TL_BEG, i + 3, 0))
        else:
            prev = col(i - 1)
            style_formula(ws.cell(DEBT_ROW_TL_BEG, i + 3, f"={prev}{DEBT_ROW_TL_END}"))
        style_input(ws.cell(DEBT_ROW_TL_AMORT, i + 3, 0))
        if not _is_proj(years, i):
            style_formula(ws.cell(DEBT_ROW_TL_PREPAY, i + 3, 0))
        else:
            style_link(ws.cell(DEBT_ROW_TL_PREPAY, i + 3, f"={cur}{DEBT_ROW_SW_TL}"))
        style_subtotal(
            ws.cell(DEBT_ROW_TL_END, i + 3,
                    f"={cur}{DEBT_ROW_TL_BEG}-{cur}{DEBT_ROW_TL_AMORT}-{cur}{DEBT_ROW_TL_PREPAY}")
        )
        style_formula(
            ws.cell(DEBT_ROW_TL_AVG, i + 3,
                    f"=({cur}{DEBT_ROW_TL_BEG}+{cur}{DEBT_ROW_TL_END})/2")
        )
        style_input(ws.cell(DEBT_ROW_TL_RATE, i + 3, 0.07), num_format=PCT)
        style_formula(
            ws.cell(DEBT_ROW_TL_INT, i + 3,
                    f"={cur}{DEBT_ROW_TL_AVG}*{cur}{DEBT_ROW_TL_RATE}")
        )

    style_label(ws.cell(DEBT_ROW_SN_TITLE, 2, "Senior Notes"), bold=True)
    style_label(ws.cell(DEBT_ROW_SN_BEG,   2, "Beginning Balance"))
    style_label(ws.cell(DEBT_ROW_SN_REPAY, 2, "Repayment at Maturity"))
    style_label(ws.cell(DEBT_ROW_SN_END,   2, "Ending Balance"))
    style_label(ws.cell(DEBT_ROW_SN_RATE,  2, "Interest Rate %"))
    style_label(ws.cell(DEBT_ROW_SN_INT,   2, "Interest Expense"))
    for i in range(n):
        cur = col(i)
        if i == 0:
            style_input(ws.cell(DEBT_ROW_SN_BEG, i + 3, 0))
        else:
            prev = col(i - 1)
            style_formula(ws.cell(DEBT_ROW_SN_BEG, i + 3, f"={prev}{DEBT_ROW_SN_END}"))
        style_input(ws.cell(DEBT_ROW_SN_REPAY, i + 3, 0))
        style_subtotal(
            ws.cell(DEBT_ROW_SN_END, i + 3,
                    f"={cur}{DEBT_ROW_SN_BEG}-{cur}{DEBT_ROW_SN_REPAY}")
        )
        style_input(ws.cell(DEBT_ROW_SN_RATE, i + 3, 0.05), num_format=PCT)
        style_formula(
            ws.cell(DEBT_ROW_SN_INT, i + 3,
                    f"={cur}{DEBT_ROW_SN_BEG}*{cur}{DEBT_ROW_SN_RATE}")
        )

    style_label(ws.cell(DEBT_ROW_TOT_TITLE, 2, "Totals"), bold=True)
    style_label(ws.cell(DEBT_ROW_TOT_BEG,   2, "Total Beginning Debt"))
    style_label(ws.cell(DEBT_ROW_TOT_END,   2, "Total Ending Debt"))
    style_label(ws.cell(DEBT_ROW_TOT_INT,   2, "Total Interest Expense"))
    style_label(ws.cell(DEBT_ROW_TOT_AMORT, 2, "Total Mandatory Amortization"))
    style_label(ws.cell(DEBT_ROW_TOT_PREP,  2, "Total Optional Prepayment"))
    for i in range(n):
        cur = col(i)
        style_subtotal(
            ws.cell(DEBT_ROW_TOT_BEG, i + 3,
                    f"={cur}{DEBT_ROW_REV_BEG}+{cur}{DEBT_ROW_TL_BEG}+{cur}{DEBT_ROW_SN_BEG}")
        )
        style_subtotal(
            ws.cell(DEBT_ROW_TOT_END, i + 3,
                    f"={cur}{DEBT_ROW_REV_END}+{cur}{DEBT_ROW_TL_END}+{cur}{DEBT_ROW_SN_END}")
        )
        style_subtotal(
            ws.cell(DEBT_ROW_TOT_INT, i + 3,
                    f"={cur}{DEBT_ROW_REV_INT}+{cur}{DEBT_ROW_TL_INT}+{cur}{DEBT_ROW_SN_INT}")
        )
        style_formula(
            ws.cell(DEBT_ROW_TOT_AMORT, i + 3, f"={cur}{DEBT_ROW_TL_AMORT}")
        )
        style_formula(
            ws.cell(DEBT_ROW_TOT_PREP, i + 3, f"={cur}{DEBT_ROW_TL_PREPAY}")
        )

    style_label(ws.cell(DEBT_ROW_SW_TITLE, 2, "Cash Sweep"), bold=True)
    style_label(ws.cell(DEBT_ROW_SW_AVAIL, 2, "Available Cash for Sweep"))
    style_label(ws.cell(DEBT_ROW_SW_APPLY, 2, "Cash Sweep Applied"))
    style_label(ws.cell(DEBT_ROW_SW_TL,    2, "Allocated to Term Loan"))
    for i in range(n):
        cur = col(i)
        if not _is_proj(years, i):
            style_formula(ws.cell(DEBT_ROW_SW_AVAIL, i + 3, 0))
            style_formula(ws.cell(DEBT_ROW_SW_APPLY, i + 3, 0))
            style_formula(ws.cell(DEBT_ROW_SW_TL,    i + 3, 0))
        else:
            style_formula(
                ws.cell(DEBT_ROW_SW_AVAIL, i + 3,
                        f"=MAX(0,CF!{cur}{CF_ROW_ENDCASH}-inp_min_cash)")
            )
            style_formula(
                ws.cell(DEBT_ROW_SW_APPLY, i + 3,
                        f"={cur}{DEBT_ROW_SW_AVAIL}*inp_cash_sweep_pct")
            )
            style_formula(
                ws.cell(DEBT_ROW_SW_TL, i + 3, f"={cur}{DEBT_ROW_SW_APPLY}")
            )

    ws.freeze_panes = "C5"
    _set_widths(ws, n)


def build_valuation(ws, years):
    apply_sheet_defaults(ws, "Valuation")
    n = len(years)
    _set_year_headers(ws, years, VAL_ROW_HEADER, projection_only=True)

    proj_indices = [i for i, (_y, lbl) in enumerate(years) if lbl.endswith("E")]
    style_label(ws.cell(VAL_ROW_EBITDA, 2, "Year-end EBITDA"))
    style_label(ws.cell(VAL_ROW_MULT,   2, "Exit EBITDA Multiple"))
    style_label(ws.cell(VAL_ROW_EV,     2, "Implied Enterprise Value"))
    style_label(ws.cell(VAL_ROW_DEBT,   2, "(Less) Year-end Total Debt"))
    style_label(ws.cell(VAL_ROW_CASH,   2, "Plus: Year-end Cash"))
    style_label(ws.cell(VAL_ROW_NCI,    2, "(Less) Minority Interest"))
    style_label(ws.cell(VAL_ROW_EQINV,  2, "Plus: Equity Investments"))
    style_label(ws.cell(VAL_ROW_EQUITY, 2, "Implied Equity Value"))
    style_label(ws.cell(VAL_ROW_SHARES, 2, "/ Diluted Shares Outstanding"))
    style_label(ws.cell(VAL_ROW_PPS,    2, "Implied Price per Share"))
    style_label(ws.cell(VAL_ROW_IRR,    2, "Implied IRR from Today"))
    style_label(ws.cell(VAL_ROW_MOIC,   2, "Implied MOIC"))

    for k, i in enumerate(proj_indices):
        cur = col(i)
        style_link(ws.cell(VAL_ROW_EBITDA, i + 3, f"=IS!{cur}{IS_ROW_EBITDA}"))
        style_link(ws.cell(VAL_ROW_MULT, i + 3, f"=Inputs!{cur}${DRV_EXIT}"), num_format=MULT)
        style_subtotal(ws.cell(VAL_ROW_EV, i + 3, f"={cur}{VAL_ROW_EBITDA}*{cur}{VAL_ROW_MULT}"))
        style_link(ws.cell(VAL_ROW_DEBT, i + 3, f"=-Debt!{cur}{DEBT_ROW_TOT_END}"))
        style_link(ws.cell(VAL_ROW_CASH, i + 3, f"=CF!{cur}{CF_ROW_ENDCASH}"))
        style_link(ws.cell(VAL_ROW_NCI, i + 3, "=-inp_minority_interest"))
        style_link(ws.cell(VAL_ROW_EQINV, i + 3, "=inp_equity_investments"))
        style_subtotal(
            ws.cell(VAL_ROW_EQUITY, i + 3,
                    f"={cur}{VAL_ROW_EV}+{cur}{VAL_ROW_DEBT}+{cur}{VAL_ROW_CASH}"
                    f"+{cur}{VAL_ROW_NCI}+{cur}{VAL_ROW_EQINV}")
        )
        style_link(ws.cell(VAL_ROW_SHARES, i + 3, f"=IS!{cur}{IS_ROW_SHARES}"))
        style_subtotal(
            ws.cell(VAL_ROW_PPS, i + 3,
                    f"=IFERROR({cur}{VAL_ROW_EQUITY}/{cur}{VAL_ROW_SHARES},0)"),
            num_format=NUM_DEC,
        )
        year_number = k + 1
        style_formula(
            ws.cell(VAL_ROW_IRR, i + 3,
                    f"=IFERROR(({cur}{VAL_ROW_PPS}/inp_current_price)^(1/{year_number})-1,0)"),
            num_format=PCT,
        )
        style_formula(
            ws.cell(VAL_ROW_MOIC, i + 3,
                    f"=IFERROR({cur}{VAL_ROW_PPS}/inp_current_price,0)"),
            num_format=MULT,
        )

    ws.freeze_panes = "C5"
    _set_widths(ws, n)


def build_summary(ws, years):
    apply_sheet_defaults(ws, "Summary")
    n = len(years)

    style_label(ws.cell(4, 2, "Company Name"), bold=True)
    style_link(ws.cell(4, 3, "=inp_company_name"), num_format=TEXT)
    style_label(ws.cell(5, 2, "Ticker"), bold=True)
    style_link(ws.cell(5, 3, "=inp_ticker"), num_format=TEXT)
    style_label(ws.cell(6, 2, "Sector"), bold=True)
    style_link(ws.cell(6, 3, "=inp_sector"), num_format=TEXT)

    style_label(ws.cell(8, 2, "Financial Summary"), bold=True)
    for i, (_yr, lbl) in enumerate(years):
        style_header(ws.cell(9, i + 3, lbl))

    style_label(ws.cell(10, 2, "Revenue"))
    style_label(ws.cell(11, 2, "Revenue Growth %"))
    style_label(ws.cell(12, 2, "EBITDA"))
    style_label(ws.cell(13, 2, "EBITDA Margin %"))
    style_label(ws.cell(14, 2, "Levered Free Cash Flow"))
    style_label(ws.cell(15, 2, "LFCF Margin %"))
    style_label(ws.cell(16, 2, "Net Debt"))
    style_label(ws.cell(17, 2, "Net Leverage (Net Debt / EBITDA)"))

    for i in range(n):
        cur = col(i)
        style_link(ws.cell(10, i + 3, f"=IS!{cur}{IS_ROW_REV}"))
        style_link(ws.cell(11, i + 3, f"=IS!{cur}{IS_ROW_GROWTH}"), num_format=PCT)
        style_link(ws.cell(12, i + 3, f"=IS!{cur}{IS_ROW_EBITDA}"))
        style_link(ws.cell(13, i + 3, f"=IS!{cur}{IS_ROW_EBITDAM}"), num_format=PCT)
        style_link(ws.cell(14, i + 3, f"=CF!{cur}{CF_ROW_LFCF}"))
        style_formula(ws.cell(15, i + 3, f"=IFERROR({cur}14/{cur}10,0)"), num_format=PCT)
        style_formula(
            ws.cell(16, i + 3, f"=Debt!{cur}{DEBT_ROW_TOT_END}-CF!{cur}{CF_ROW_ENDCASH}")
        )
        style_subtotal(
            ws.cell(17, i + 3, f"=IFERROR({cur}16/{cur}12,0)"),
            num_format=MULT,
        )

    style_label(ws.cell(19, 2, "Returns Table"), bold=True)
    for j, h in enumerate(["Year", "Implied Price per Share", "Implied IRR", "Implied MOIC"]):
        style_header(ws.cell(20, j + 2, h))

    proj_indices = [i for i, (_y, lbl) in enumerate(years) if lbl.endswith("E")]
    for k, (yr_off, label) in enumerate([(1, "Y1"), (3, "Y3"), (5, "Y5"), (10, "Y10")]):
        r = 21 + k
        i_col = proj_indices[yr_off - 1]
        cur = col(i_col)
        style_label(ws.cell(r, 2, f"{label} ({years[i_col][1]})"), bold=True)
        style_link(ws.cell(r, 3, f"=Valuation!{cur}{VAL_ROW_PPS}"), num_format=NUM_DEC)
        style_link(ws.cell(r, 4, f"=Valuation!{cur}{VAL_ROW_IRR}"), num_format=PCT)
        style_link(ws.cell(r, 5, f"=Valuation!{cur}{VAL_ROW_MOIC}"), num_format=MULT)

    ws.freeze_panes = "C5"
    _set_widths(ws, n)


def build_sensitivity(ws, _years):
    apply_sheet_defaults(ws, "Sensitivity")

    style_label(ws.cell(4, 2, "Target Year Offset (1 = first projection year)"))
    style_input(ws.cell(4, 3, 5), num_format="0")  # named sens_target_year_offset

    style_label(ws.cell(6, 2, "Exit Multiple"), bold=True)
    header_b = ws.cell(6, 3, '="Implied Price per Share at Year " & sens_target_year_offset')
    header_b.font = LABEL_BOLD_FONT

    multiples = [6, 7, 8, 9, 10, 11, 12, 13, 14]
    # IS EBITDA row, Debt total ending row, CF ending cash row, IS shares row.
    # Year columns on those tabs run from F to O (cols 6..15) for the 10 projection years.
    is_ebitda_range  = f"IS!$F${IS_ROW_EBITDA}:$O${IS_ROW_EBITDA}"
    debt_end_range   = f"Debt!$F${DEBT_ROW_TOT_END}:$O${DEBT_ROW_TOT_END}"
    cf_endcash_range = f"CF!$F${CF_ROW_ENDCASH}:$O${CF_ROW_ENDCASH}"
    is_shares_range  = f"IS!$F${IS_ROW_SHARES}:$O${IS_ROW_SHARES}"

    for k, m in enumerate(multiples):
        r = 7 + k
        style_input(ws.cell(r, 2, m), num_format=MULT)
        formula = (
            f"=(INDEX({is_ebitda_range},sens_target_year_offset)*$B{r}"
            f"-INDEX({debt_end_range},sens_target_year_offset)"
            f"+INDEX({cf_endcash_range},sens_target_year_offset)"
            f"-inp_minority_interest+inp_equity_investments)"
            f"/INDEX({is_shares_range},sens_target_year_offset)"
        )
        style_formula(ws.cell(r, 3, formula), num_format=NUM_DEC)

    style_label(ws.cell(17, 2, "2D Sensitivity (Exit Multiple x Year)"), bold=True)
    for j, h in enumerate(["Multiple", "Y3", "Y5", "Y7", "Y10"]):
        style_header(ws.cell(18, j + 2, h))

    year_offsets = [3, 5, 7, 10]
    for k, m in enumerate(multiples):
        r = 19 + k
        style_input(ws.cell(r, 2, m), num_format=MULT)
        for j, off in enumerate(year_offsets):
            formula = (
                f"=(INDEX({is_ebitda_range},{off})*$B{r}"
                f"-INDEX({debt_end_range},{off})"
                f"+INDEX({cf_endcash_range},{off})"
                f"-inp_minority_interest+inp_equity_investments)"
                f"/INDEX({is_shares_range},{off})"
            )
            style_formula(ws.cell(r, j + 3, formula), num_format=NUM_DEC)

    ws.column_dimensions["B"].width = 38
    for letter in ["C", "D", "E", "F"]:
        ws.column_dimensions[letter].width = 20


def build_capiq_data(ws):
    """Mirror of capiq_fetcher.xlsx Fetcher tab — values only, no CapIQ formulas.

    Layout sourced from shared.capiq_layout so this stays in lockstep with
    the fetcher. shared/fetch_capiq.py writes hardcoded values into the data
    cells; the cells are left empty here.
    """
    apply_sheet_defaults(ws, "CapIQ Data")
    ws.sheet_state = "hidden"
    ws.sheet_properties.tabColor = "808080"

    banner = ws.cell(capiq_layout.ROW_BANNER, 2,
                     "_CapIQ_Data — DO NOT EDIT MANUALLY. Populated by shared/fetch_capiq.py.")
    banner.font = BANNER_FONT

    style_label(ws.cell(capiq_layout.ROW_RUN_VIA, 2, "Run via:"), bold=True)
    style_label(ws.cell(capiq_layout.ROW_RUN_VIA, 3, "python -m shared.fetch_capiq <TICKER>"))
    style_label(ws.cell(capiq_layout.ROW_TICKER, 2, "Ticker:"), bold=True)
    style_label(ws.cell(capiq_layout.ROW_DATE, 2, "Date"), bold=True)
    style_label(ws.cell(capiq_layout.ROW_FETCHER_DATE, 2, "Fetcher Run-Date"), bold=True)

    for j, h in enumerate(capiq_layout.COL_HEADERS):
        style_header(ws.cell(capiq_layout.ROW_COL_HEADERS, j + 2, h))

    # Section A: metadata labels in B; values in F (left blank — fetch script writes them).
    for r, label, _ in capiq_layout.METADATA:
        style_label(ws.cell(r, 2, label))

    # Section B: current state. Two rows have formulas (Market Cap, Enterprise Value).
    for r, label, is_calc in capiq_layout.CURRENT_STATE:
        if is_calc:
            style_label(ws.cell(r, 2, label), bold=True)
        else:
            style_label(ws.cell(r, 2, label))
    # Market Cap = F18 * F19
    style_subtotal(ws.cell(20, 6, "=F18*F19"), num_format=NUM)
    # Enterprise Value = F20 - F21 - F22 + F23 + F24 + F25 - F26 - F27
    style_subtotal(
        ws.cell(28, 6, "=F20-F21-F22+F23+F24+F25-F26-F27"),
        num_format=NUM,
    )

    # Section C: historicals — labels only; data in C/D/E populated by fetch.
    for r, label in capiq_layout.HISTORICALS:
        style_label(ws.cell(r, 2, label))

    ws.column_dimensions["B"].width = 35
    for letter in ["C", "D", "E", "F"]:
        ws.column_dimensions[letter].width = 14


def build_broker_data(ws):
    """Mirror of broker_fetcher.xlsx Fetcher tab.

    P&L grid (rows 13-20 cols C-H) and sentiment fetched values (B31, B32, B34)
    are populated by shared/fetch_broker_estimates.py at runtime. Implied
    growth/margin formulas (rows 23-28) and implied upside (row 33) live in
    this workbook so they update when historicals refresh.
    """
    apply_sheet_defaults(ws, "Broker Data")
    ws.sheet_state = "hidden"
    ws.sheet_properties.tabColor = "808080"

    banner = ws.cell(broker_layout.ROW_BANNER, 2,
                     "_Broker_Data — DO NOT EDIT MANUALLY. Populated by shared/fetch_broker_estimates.py.")
    banner.font = BANNER_FONT

    style_label(ws.cell(broker_layout.ROW_LAST_FETCH, 2, "Last fetch:"), bold=True)
    style_label(ws.cell(broker_layout.ROW_TICKER, 2, "Ticker:"), bold=True)
    style_label(ws.cell(broker_layout.ROW_FY1_YEAR, 2, "FY1 fiscal year:"), bold=True)
    style_label(ws.cell(broker_layout.ROW_FY2_YEAR, 2, "FY2 fiscal year:"), bold=True)
    style_label(ws.cell(broker_layout.ROW_FY3_YEAR, 2, "FY3 fiscal year:"), bold=True)

    for j, h in enumerate(broker_layout.COL_HEADERS):
        style_header(ws.cell(broker_layout.ROW_COL_HEADERS, j + 2, h))

    for r, label, *_ in broker_layout.PNL:
        style_label(ws.cell(r, 2, label))

    for r, label, formula in broker_layout.IMPLIED:
        style_label(ws.cell(r, 2, label))
        c = ws.cell(r, 3, formula)
        style_formula(c, num_format=PCT)

    for r, label, _ in broker_layout.SENTIMENT:
        style_label(ws.cell(r, 2, label))
    for r, formula in broker_layout.SENTIMENT_FORMULAS_IN_MODEL.items():
        c = ws.cell(r, 3, formula)
        style_formula(c, num_format=PCT)

    ws.column_dimensions["B"].width = 38
    for letter in ["C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[letter].width = 14


def build():
    wb = Workbook()

    # Pre-enable iterative calc so the interest <-> taxes <-> cash sweep
    # circularity converges when the workbook is opened.
    wb.calculation.iterate = True
    wb.calculation.iterateCount = 100
    wb.calculation.iterateDelta = 0.001

    years = make_year_columns()

    wb.remove(wb.active)
    cover_ws = wb.create_sheet("Cover")
    inputs_ws = wb.create_sheet("Inputs")
    is_ws = wb.create_sheet("IS")
    cf_ws = wb.create_sheet("CF")
    debt_ws = wb.create_sheet("Debt")
    val_ws = wb.create_sheet("Valuation")
    summary_ws = wb.create_sheet("Summary")
    sens_ws = wb.create_sheet("Sensitivity")
    capiq_ws = wb.create_sheet("_CapIQ_Data")
    broker_ws = wb.create_sheet("_Broker_Data")

    # Inputs first so the named ranges it registers are available to the others.
    build_inputs(inputs_ws, years)
    register_inputs_named_ranges(wb)
    register_driver_named_ranges(wb)

    build_cover(cover_ws, years)
    build_is(is_ws, years)
    build_cf(cf_ws, years)
    build_debt(debt_ws, years)
    build_valuation(val_ws, years)
    build_summary(summary_ws, years)
    build_sensitivity(sens_ws, years)
    build_capiq_data(capiq_ws)
    build_broker_data(broker_ws)

    add_named_range(wb, "sens_target_year_offset", "Sensitivity", "$C$4")

    TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(TEMPLATE_PATH)
    return TEMPLATE_PATH


if __name__ == "__main__":
    out = build()
    print(f"Wrote {out}")
