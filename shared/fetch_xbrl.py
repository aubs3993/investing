"""Fill a ticker's _CapIQ_Data tab from SEC XBRL + Yahoo Finance (no CapIQ).

Usage:
    python -m shared.fetch_xbrl <TICKER>
    python -m shared.fetch_xbrl <TICKER> --price 123.45
    python -m shared.fetch_xbrl <TICKER> --offline
    python -m shared.fetch_xbrl <TICKER> --model-path PATH

Drop-in replacement for shared.fetch_capiq (which needs live Excel + the
CapIQ plugin): writes the same _CapIQ_Data layout (shared/capiq_layout.py),
so extract_historicals, the IS/CF/Inputs links, and populate_drivers need no
changes. Runs anywhere with internet — no Excel required.

Data sources:
  - SEC EDGAR companyfacts JSON (all financials). Requires SEC_USER_AGENT in
    .env, formatted "Name email@example.com" (SEC rejects anonymous clients).
  - SEC EDGAR submissions JSON (sector = SIC description, filer category).
  - Yahoo Finance v8 chart endpoint (current price; browser UA, no key).
  Raw downloads are cached next to the model file and reused on network
  failure (or with --offline).

Design ported from merlin_stock_updates (stock_updates/fundamentals.py,
sec.py, quotes.py), adapted for 3-fiscal-year series extraction:
  - Canonical FY periods come from OperatingIncomeLoss/Revenue duration facts
    (340-380 day windows — 52/53-week retail years pass), 10-K forms only so
    proxy-statement re-tags can't pollute, deduped by end-date proximity with
    latest-filed winning (restated comparatives supersede originals).
    Company-facts fy/fp fields are ignored — they stamp the filing's fiscal
    context, not the fact's.
  - Every line item is then resolved per canonical period across a tag-fallback
    chain (per-period merge, not merlin's winner-take-all tag rule, so tag
    migrations mid-history can't drop older years).
  - Absent is blank + console note; only a genuinely tagged zero is 0.
    Exception: restructuring/impairment add-backs, where an untagged period is
    a genuine $0 (merlin convention).

Conventions (differences vs the CapIQ fetcher worth knowing):
  - Units: $ millions (shares too); DPS and price are per-share dollars.
  - EBITDA row is ADJUSTED EBITDA = EBIT + D&A + restructuring + impairments
    (merlin's standardized-EBITDA add-backs; SBC deliberately NOT added back —
    CapIQ convention). The add-back uses max(combined tag, sum of individual
    tags) per period: filers tag the combined "restructuring and impairment"
    form OR the individual forms, and max() can neither double-count nor miss
    a solo combined tag (merlin skips combined tags, fine for its LTM window,
    wrong for e.g. LULU FY2023 where $74.5M lives only in the combined tag).
  - EBIT row = adj EBITDA - D&A and Total Opex row = Gross Profit - adj
    EBITDA, so the IS tab's subtotal identities (EBITDA = GP - Opex,
    EBIT = EBITDA - D&A) hold exactly on an adjusted basis.
  - Total Debt includes finance leases, excludes operating leases (CapIQ
    convention; material for retailers).
  - Shares outstanding = dei cover-page count (basic); margins are as-reported
    GAAP presentation, not CapIQ-standardized (COGS keeps embedded D&A).
"""
from __future__ import annotations

import argparse
import json
import os
import time
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import quote as urlquote

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

from shared import capiq_layout
from shared.excel_helpers import validate_field_labels
from shared.model_path import resolve_model_path
from shared.tickers import fs_ticker, validate_ticker

REPO_ROOT = Path(__file__).resolve().parent.parent
load_dotenv(REPO_ROOT / ".env")

# --- SEC endpoints and pacing (merlin sec.py / config.py) ---
SEC_TICKER_MAP_URL = "https://www.sec.gov/files/company_tickers.json"
SEC_SUBMISSIONS_URL = "https://data.sec.gov/submissions/CIK{cik:010d}.json"
SEC_COMPANYFACTS_URL = "https://data.sec.gov/api/xbrl/companyfacts/CIK{cik:010d}.json"
SEC_REQUEST_INTERVAL_SECONDS = 0.12   # ~8.3 req/s, under SEC's 10/s cap
SEC_MAX_RETRIES = 3
SEC_TIMEOUT_SECONDS = 30

# --- Yahoo (merlin quotes.py; chart endpoint only — no crumb needed) ---
YAHOO_CHART_URL = "https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
YAHOO_REQUEST_INTERVAL_SECONDS = 0.25
YAHOO_MAX_RETRIES = 2
YAHOO_TIMEOUT_SECONDS = 20
# Yahoo rejects default python-requests UAs; a plain browser UA is expected.
YAHOO_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
)

GAAP = "us-gaap"
DEI = "dei"
MM = 1e6

# Annual duration window: 52-week retail years span 363 days end-minus-start,
# 53-week years 370 — both inside [340, 380] (merlin fundamentals.py).
ANNUAL_MIN_DAYS, ANNUAL_MAX_DAYS = 340, 380
# Annual durations only come from annual reports; excluding other forms keeps
# proxy statements (LULU's DEFC14A re-tags NetIncomeLoss) out of the series.
ANNUAL_FORMS = {"10-K", "10-K/A", "10-KT", "10-KT/A"}
INSTANT_FORMS = {"10-K", "10-K/A", "10-Q", "10-Q/A"}
# Two "annual" periods whose ends are this close are the same fiscal year
# tagged with a start date shifted by a day between filings.
END_MERGE_DAYS = 20
# Canonical periods should be ~1 year apart; outside this band means a fiscal
# year change or missing year slipped into the series.
SPACING_MIN_DAYS, SPACING_MAX_DAYS = 330, 400
# An instant older than this vs. the cash balance-sheet date is a tag the
# filer no longer reports (LULU's MinorityInterest facts end in 2012), not a
# current value.
STALE_DAYS = 550
DEBT_FRESH_DAYS = 100
SHARES_FRESH_DAYS = 400

# --- Tag chains: first tag with data for the period wins ---
# (income-statement chains authored here; balance-sheet chains from merlin)
REVENUE_TAGS = (
    "RevenueFromContractWithCustomerExcludingAssessedTax",
    "RevenueFromContractWithCustomerIncludingAssessedTax",
    "Revenues",
    "SalesRevenueNet",
    "SalesRevenueGoodsNet",
)
COGS_TAGS = ("CostOfGoodsAndServicesSold", "CostOfRevenue", "CostOfGoodsSold",
             "CostOfSales")
GROSS_PROFIT_TAGS = ("GrossProfit",)
OI_TAGS = ("OperatingIncomeLoss",)
DA_TAGS = (
    "DepreciationDepletionAndAmortization",
    "DepreciationAmortizationAndAccretionNet",
    "DepreciationAndAmortization",
)
CAPEX_TAGS = ("PaymentsToAcquirePropertyPlantAndEquipment",
              "PaymentsToAcquireProductiveAssets")
SBC_TAGS = ("ShareBasedCompensation",)
DPS_TAGS = ("CommonStockDividendsPerShareDeclared",
            "CommonStockDividendsPerShareCashPaid")

# Adjusted-EBITDA add-backs (merlin fundamentals.py:88-112). SBC deliberately
# excluded. Individual concepts are summed; the combined tag competes via
# max() (see module docstring).
RESTRUCTURING_TAGS = ("RestructuringCharges", "RestructuringCosts")
IMPAIRMENT_TAG_GROUPS = (
    ("GoodwillImpairmentLoss", "GoodwillAndIntangibleAssetImpairment"),
    ("AssetImpairmentCharges", "ImpairmentOfLongLivedAssetsHeldForUse",
     "TangibleAssetImpairmentCharges"),
    ("ImpairmentOfIntangibleAssetsExcludingGoodwill",
     "ImpairmentOfIntangibleAssetsIndefinitelivedExcludingGoodwill",
     "ImpairmentOfIntangibleAssetsFinitelived"),
)
COMBINED_RESTR_IMPAIR_TAGS = (
    "RestructuringCostsAndAssetImpairmentCharges",
    "RestructuringSettlementAndImpairmentProvisions",
)

# Balance-sheet chains (merlin fundamentals.py:61-86, 180-192).
CASH_TAGS = (
    "CashAndCashEquivalentsAtCarryingValue",
    "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents",
)
STI_TAGS = ("ShortTermInvestments", "MarketableSecuritiesCurrent",
            "OtherShortTermInvestments")
# Noncurrent only — the current-marketable-securities concept belongs to the
# ST Investments row; splitting this way keeps the EV formula (which subtracts
# both rows) from double-counting.
MARKETABLE_SEC_TAGS = ("MarketableSecuritiesNoncurrent",)
PREFERRED_TAGS = ("PreferredStockValue", "PreferredStockValueOutstanding")
MINORITY_TAGS = ("MinorityInterest",)
# Three non-overlapping equity-stake concepts, summed. Never add aggregate
# tags (LongTermInvestments/Investments) — they contain these components.
EQUITY_INV_TAG_GROUPS = (
    ("EquityMethodInvestments",),
    ("EquitySecuritiesFvNiNoncurrent", "EquitySecuritiesFvNiCurrentAndNoncurrent",
     "EquitySecuritiesFvNi"),
    ("EquitySecuritiesWithoutReadilyDeterminableFairValueAmount",),
)
DEBT_AGGREGATE_TAGS = ("DebtAndCapitalLeaseObligations",
                       "DebtLongtermAndShorttermCombinedAmount")
DEBT_LT_TAGS = ("LongTermDebtNoncurrent", "LongTermDebtAndCapitalLeaseObligations",
                "LongTermDebt")
DEBT_CURRENT_TAGS = ("LongTermDebtCurrent",
                     "LongTermDebtAndCapitalLeaseObligationsCurrent")
DEBT_ST_TAGS = ("ShortTermBorrowings", "CommercialPaper")
FINANCE_LEASE_TAGS = ("FinanceLeaseLiabilityCurrent", "FinanceLeaseLiabilityNoncurrent")


# --- HTTP sessions -----------------------------------------------------------
class SecSession:
    """Rate-limited EDGAR session (merlin sec.py, plus 5xx/connection retries)."""

    def __init__(self, user_agent: str):
        self._session = requests.Session()
        self._session.headers["User-Agent"] = user_agent
        self._last_request = 0.0

    def get_json(self, url):
        last_err = None
        for attempt in range(SEC_MAX_RETRIES + 1):
            self._pace()
            try:
                resp = self._session.get(url, timeout=SEC_TIMEOUT_SECONDS)
            except requests.RequestException as e:
                last_err = e
                if attempt < SEC_MAX_RETRIES:
                    time.sleep(2 ** attempt)
                    continue
                raise
            if resp.status_code in (429, 500, 502, 503, 504) or (
                # SEC also serves 403 on rate-threshold blocks, not just bad
                # User-Agents — retry once before blaming the UA.
                resp.status_code == 403 and attempt == 0
            ):
                if attempt < SEC_MAX_RETRIES:
                    time.sleep(2 ** attempt)
                    continue
            if resp.status_code == 403:
                raise SystemExit(
                    f"SEC returned 403 for {url}. Either SEC_USER_AGENT is "
                    f"missing/malformed in .env (must look like "
                    f"'Name email@example.com') or EDGAR rate-blocked this IP."
                )
            resp.raise_for_status()
            return resp.json()
        raise RuntimeError(f"SEC request kept failing: {url} ({last_err})")

    def _pace(self):
        elapsed = time.monotonic() - self._last_request
        if elapsed < SEC_REQUEST_INTERVAL_SECONDS:
            time.sleep(SEC_REQUEST_INTERVAL_SECONDS - elapsed)
        self._last_request = time.monotonic()


def yahoo_last_close(symbol: str) -> dict:
    """Last daily close for one symbol via the v8 chart endpoint.

    Returns {"price", "as_of", "name"}. Raises on any failure — the caller
    degrades to a blank price with a console warning.
    """
    session = requests.Session()
    session.headers["User-Agent"] = YAHOO_USER_AGENT
    session.headers["Accept"] = "application/json,text/plain,*/*"
    resp = None
    for attempt in range(YAHOO_MAX_RETRIES + 1):
        resp = session.get(
            YAHOO_CHART_URL.format(symbol=urlquote(symbol)),
            params={"range": "5d", "interval": "1d"},
            timeout=YAHOO_TIMEOUT_SECONDS,
        )
        if resp.status_code == 429 and attempt < YAHOO_MAX_RETRIES:
            time.sleep(2 ** attempt)
            continue
        break
    data = resp.json()
    chart = data.get("chart") or {}
    results = chart.get("result") or []
    if not results:
        error = chart.get("error") or {}
        raise RuntimeError(error.get("description") or error.get("code")
                           or f"no chart data (HTTP {resp.status_code})")
    result = results[0]
    meta = result.get("meta") or {}
    closes = list(zip(
        result.get("timestamp") or [],
        (((result.get("indicators") or {}).get("quote") or [{}])[0]
         .get("close") or []),
    ))
    closes = [(ts, c) for ts, c in closes
              if isinstance(c, (int, float)) and not isinstance(c, bool)]
    if not closes:
        raise RuntimeError("chart returned no closes")
    last_ts, price = closes[-1]
    return {
        "price": float(price),
        "as_of": datetime.fromtimestamp(last_ts, timezone.utc).date().isoformat(),
        "name": meta.get("longName") or meta.get("shortName") or "",
    }


def _name_matches(entity_name: str, yahoo_name: str) -> bool:
    """Loose guard that Yahoo resolved the symbol to the intended company
    (a recycled ticker pricing the wrong company is worse than a blank)."""
    if not yahoo_name:
        return True
    for word in (entity_name or "").lower().replace(",", " ").replace(".", " ").split():
        if word not in ("the",):
            return word in yahoo_name.lower()
    return True


# --- Cached fetches ----------------------------------------------------------
def _fetch_cached(fetch, cache_path: Path, offline: bool, what: str,
                  required: bool = True):
    """Fetch JSON via `fetch()`, mirroring to cache_path; fall back to the
    cache on network failure (always used with --offline)."""
    if offline:
        if cache_path.exists():
            return json.loads(cache_path.read_text(encoding="utf-8"))
        if required:
            raise SystemExit(f"--offline but no cache at {cache_path}")
        return None
    try:
        data = fetch()
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        cache_path.write_text(json.dumps(data), encoding="utf-8")
        return data
    except (requests.RequestException, RuntimeError, SystemExit):
        if cache_path.exists():
            print(f"  WARNING: {what} fetch failed; using cached copy "
                  f"{cache_path.name}")
            return json.loads(cache_path.read_text(encoding="utf-8"))
        if required:
            raise
        print(f"  WARNING: {what} fetch failed and no cache — skipping")
        return None


# --- Fact resolution ---------------------------------------------------------
def _iso(d: str) -> date:
    return date.fromisoformat(d)


def _facts_list(facts, taxonomy, tag, unit):
    return (facts.get("facts", {}).get(taxonomy, {}).get(tag, {})
            .get("units", {}).get(unit, []))


def canonical_annual_periods(facts, notes: list[str], n: int = 3):
    """The n most recent complete fiscal years as [(start, end)], oldest first.

    Anchored on OperatingIncomeLoss/Revenue durations from 10-K forms.
    Same-year duplicates (end dates within END_MERGE_DAYS, from start dates
    shifted a day between filings) collapse to the latest-filed tagging.
    """
    candidates = {}  # (start, end) -> filed
    for tag in OI_TAGS + REVENUE_TAGS:
        for it in _facts_list(facts, GAAP, tag, "USD"):
            if "start" not in it or it.get("form") not in ANNUAL_FORMS:
                continue
            start, end = _iso(it["start"]), _iso(it["end"])
            if not ANNUAL_MIN_DAYS <= (end - start).days <= ANNUAL_MAX_DAYS:
                continue
            key = (start, end)
            if it.get("filed", "") > candidates.get(key, ("",))[0]:
                candidates[key] = (it.get("filed", ""),)
    if not candidates:
        raise SystemExit(
            "No annual OperatingIncomeLoss/Revenue periods found in "
            "companyfacts — non-US-GAAP filer (IFRS/20-F) or fresh IPO?"
        )
    picked = []
    for (start, end), (filed,) in sorted(
        candidates.items(), key=lambda kv: (kv[0][1], kv[1][0]), reverse=True
    ):
        if any(abs((end - p_end).days) <= END_MERGE_DAYS for _, p_end in picked):
            continue
        picked.append((start, end))
        if len(picked) == n:
            break
    picked.reverse()
    if len(picked) < n:
        notes.append(f"only {len(picked)} complete fiscal years available")
    for (_, e1), (_, e2) in zip(picked, picked[1:]):
        if not SPACING_MIN_DAYS <= (e2 - e1).days <= SPACING_MAX_DAYS:
            notes.append(
                f"fiscal periods ending {e1} and {e2} are not ~1 year apart — "
                f"fiscal-year change or missing year in between?"
            )
    return picked


def period_value(facts, tags, start: date, end: date, unit="USD",
                 taxonomy=GAAP):
    """Value of the first tag in `tags` covering exactly this fiscal period.

    End must match exactly; start may differ by <=2 days (filers shift
    comparative start dates between filings). Latest filed wins so restated
    comparatives supersede originals.
    """
    for tag in tags:
        best = None
        for it in _facts_list(facts, taxonomy, tag, unit):
            if "start" not in it or it.get("form") not in ANNUAL_FORMS:
                continue
            if it["end"] != end.isoformat():
                continue
            if abs((_iso(it["start"]) - start).days) > 2:
                continue
            if best is None or it.get("filed", "") > best.get("filed", ""):
                best = it
        if best is not None:
            return float(best["val"]), tag
    return None, None


def da_value(facts, start: date, end: date, notes: list[str], fy: str):
    """D&A for one fiscal year: combined-tag chain, else Depreciation +
    AmortizationOfIntangibleAssets (noting a possible understatement when
    only depreciation is tagged)."""
    val, tag = period_value(facts, DA_TAGS, start, end)
    if val is not None:
        return val, tag
    dep, _ = period_value(facts, ("Depreciation",), start, end)
    if dep is None:
        return None, None
    amort, _ = period_value(facts, ("AmortizationOfIntangibleAssets",), start, end)
    if amort is None:
        notes.append(f"{fy}: D&A = Depreciation only (amortization untagged) — "
                     f"may understate")
    return dep + (amort or 0.0), "Depreciation+Amortization"


def addback_value(facts, start: date, end: date, notes: list[str], fy: str):
    """Restructuring + impairment add-back for one fiscal year.

    max(sum of individual concepts, combined concept): equal when a filer tags
    both (the combined tag is their sum by definition), and max() still
    catches the charge when only one form is tagged. Untagged period = genuine
    $0 (add-backs are optional charges, unlike core line items).
    """
    individual = 0.0
    for group in (RESTRUCTURING_TAGS,) + IMPAIRMENT_TAG_GROUPS:
        val, _ = period_value(facts, group, start, end)
        individual += val or 0.0
    combined, _ = period_value(facts, COMBINED_RESTR_IMPAIR_TAGS, start, end)
    combined = combined or 0.0
    addback = max(individual, combined)
    if addback:
        source = "combined tag" if combined > individual else "individual tags"
        notes.append(f"{fy}: adj EBITDA adds back {addback / MM:,.1f}mm "
                     f"restructuring/impairment ({source})")
    return addback


def latest_instant(facts, tags, unit="USD", taxonomy=DEI, forms=INSTANT_FORMS):
    """(value, end, tag) of the newest instant across the chain; newest end
    wins, filed breaks ties (amended values supersede)."""
    best = None
    for tag in tags:
        for it in _facts_list(facts, taxonomy, tag, unit):
            if "start" in it or it.get("form") not in forms:
                continue
            rank = (it["end"], it.get("filed", ""))
            if best is None or rank > best[0]:
                best = (rank, float(it["val"]), _iso(it["end"]), tag)
    if best is None:
        return None, None, None
    return best[1], best[2], best[3]


def gaap_instant(facts, tags):
    return latest_instant(facts, tags, taxonomy=GAAP)


def fresh_instant(facts, tags, anchor: date, notes: list[str], label: str,
                  max_age_days: int = STALE_DAYS):
    """Like gaap_instant but treats values older than max_age_days vs the
    balance-sheet anchor date as no-longer-reported (None), not current."""
    val, end, tag = gaap_instant(facts, tags)
    if val is None:
        return None, None
    if anchor and (anchor - end).days > max_age_days:
        notes.append(f"{label}: last tagged {end} — stale, treated as absent")
        return None, None
    return val, tag


def total_debt(facts, anchor: date, notes: list[str]):
    """Merlin's debt cascade, compacted: fresh lease-inclusive aggregate,
    else components + finance leases. Operating leases excluded by design."""
    for tag in DEBT_AGGREGATE_TAGS:
        val, end, _ = gaap_instant(facts, (tag,))
        if val is not None and anchor and (anchor - end).days <= DEBT_FRESH_DAYS:
            notes.append(f"total debt from aggregate {tag} (leases included)")
            return val
    total, found, leases_included = 0.0, [], False
    for chain in (DEBT_LT_TAGS, DEBT_CURRENT_TAGS):
        for tag in chain:
            val, end, _ = gaap_instant(facts, (tag,))
            if val is not None and anchor and (anchor - end).days <= DEBT_FRESH_DAYS:
                total += val
                found.append(tag)
                if "CapitalLeaseObligations" in tag:
                    leases_included = True
                break
    for tag in DEBT_ST_TAGS:
        val, end, _ = gaap_instant(facts, (tag,))
        if val is not None and anchor and (anchor - end).days <= DEBT_FRESH_DAYS:
            total += val
            found.append(tag)
    if not leases_included:
        for tag in FINANCE_LEASE_TAGS:
            val, end, _ = gaap_instant(facts, (tag,))
            if val is not None and anchor and (anchor - end).days <= DEBT_FRESH_DAYS:
                total += val
                found.append(tag)
    if not found:
        notes.append("no fresh debt tags — debt-free filer? Total Debt = 0")
        return 0.0
    notes.append("total debt components: " + ", ".join(found))
    return total


def shares_outstanding(facts, notes: list[str]):
    """Cover-page share count (dei, basic) in raw shares; falls back to the
    latest annual weighted-average diluted count."""
    val, end, _ = latest_instant(facts, ("EntityCommonStockSharesOutstanding",),
                                 unit="shares", taxonomy=DEI)
    if val is not None and (date.today() - end).days <= SHARES_FRESH_DAYS:
        notes.append(f"shares = dei cover-page count as of {end} (basic, "
                     f"single-class assumption)")
        return val
    best = None
    for it in _facts_list(facts, GAAP,
                          "WeightedAverageNumberOfDilutedSharesOutstanding",
                          "shares"):
        if "start" not in it or it.get("form") not in ANNUAL_FORMS:
            continue
        start, end_d = _iso(it["start"]), _iso(it["end"])
        if not ANNUAL_MIN_DAYS <= (end_d - start).days <= ANNUAL_MAX_DAYS:
            continue
        rank = (it["end"], it.get("filed", ""))
        if best is None or rank > best[0]:
            best = (rank, float(it["val"]))
    if best is not None:
        notes.append("shares = latest annual weighted-average diluted "
                     "(no fresh cover-page count)")
        return best[1]
    notes.append("no share count found")
    return None


def _fy_label(end: date) -> str:
    """Retail convention: fiscal year named for the calendar year holding most
    of the period (LULU's FY ended 2026-02-01 is fiscal 2025)."""
    return f"FY{(end - timedelta(days=183)).year}"


# --- Extraction --------------------------------------------------------------
def extract_financials(facts, submissions, notes: list[str]) -> dict:
    """All values for the _CapIQ_Data tab, keyed by capiq_layout labels.

    Historicals come back as 3-element lists (FY-2, FY-1, FY) in $mm; None
    means untagged (rendered blank).
    """
    periods = canonical_annual_periods(facts, notes)
    hist = {label: [None] * 3 for _, label in capiq_layout.HISTORICALS}
    pad = 3 - len(periods)

    def mm(v):
        return None if v is None else v / MM

    for i, (start, end) in enumerate(periods):
        col = pad + i
        fy = _fy_label(end)

        rev, rev_tag = period_value(facts, REVENUE_TAGS, start, end)
        cogs, _ = period_value(facts, COGS_TAGS, start, end)
        gp, _ = period_value(facts, GROSS_PROFIT_TAGS, start, end)
        if gp is None and rev is not None and cogs is not None:
            gp = rev - cogs
            notes.append(f"{fy}: GrossProfit untagged — derived Revenue-COGS")
        if cogs is None and rev is not None and gp is not None:
            cogs = rev - gp
            notes.append(f"{fy}: COGS untagged — derived Revenue-GrossProfit")
        ebit_reported, _ = period_value(facts, OI_TAGS, start, end)
        da, _ = da_value(facts, start, end, notes, fy)
        capex, _ = period_value(facts, CAPEX_TAGS, start, end)
        sbc, _ = period_value(facts, SBC_TAGS, start, end)
        dps, _ = period_value(facts, DPS_TAGS, start, end, unit="USD/shares")

        adj_ebitda = adj_ebit = opex = None
        if ebit_reported is not None and da is not None:
            addback = addback_value(facts, start, end, notes, fy)
            adj_ebitda = ebit_reported + da + addback
            adj_ebit = adj_ebitda - da
            if gp is not None:
                opex = gp - adj_ebitda
                if opex < 0:
                    notes.append(f"{fy}: derived Total Opex is negative — check")
        elif ebit_reported is None:
            notes.append(f"{fy}: OperatingIncomeLoss untagged — EBITDA/EBIT/Opex blank")
        else:
            notes.append(f"{fy}: D&A untagged — EBITDA/EBIT/Opex blank")

        if rev is not None and cogs is not None and gp is not None:
            if abs((rev - cogs) - gp) > max(abs(gp) * 0.005, 1e5):
                notes.append(f"{fy}: Revenue - COGS != GrossProfit "
                             f"({(rev - cogs - gp) / MM:,.1f}mm gap) — check tags")

        hist["Revenue"][col] = mm(rev)
        hist["COGS"][col] = mm(cogs)
        hist["Gross Profit"][col] = mm(gp)
        hist["Total Opex"][col] = mm(opex)
        hist["D&A"][col] = mm(da)
        hist["EBITDA"][col] = mm(adj_ebitda)
        hist["EBIT"][col] = mm(adj_ebit)
        hist["Capex"][col] = mm(capex)
        hist["SBC"][col] = mm(sbc)
        hist["DPS"][col] = dps  # already per-share dollars

    # --- Current state (latest balance sheet) ---
    cash, cash_end, cash_tag = gaap_instant(facts, CASH_TAGS)
    if cash_tag == CASH_TAGS[1]:
        notes.append("cash includes restricted cash (only the inclusive tag "
                     "is currently filed)")
    if cash_end is None:
        notes.append("no cash balance found")
    sti, _ = fresh_instant(facts, STI_TAGS, cash_end, notes, "ST Investments")
    mkt_sec, _ = fresh_instant(facts, MARKETABLE_SEC_TAGS, cash_end, notes,
                               "Marketable Securities")
    preferred, _ = fresh_instant(facts, PREFERRED_TAGS, cash_end, notes,
                                 "Preferred Equity")
    minority, _ = fresh_instant(facts, MINORITY_TAGS, cash_end, notes,
                                "Minority Interest")
    eq_inv_total, eq_found = 0.0, False
    for group in EQUITY_INV_TAG_GROUPS:
        val, _ = fresh_instant(facts, group, cash_end, notes, "Equity Investments")
        if val is not None:
            eq_inv_total += val
            eq_found = True
    debt = total_debt(facts, cash_end, notes)
    shares = shares_outstanding(facts, notes)

    sic = (submissions or {}).get("sicDescription")
    category = (submissions or {}).get("category") or \
        (submissions or {}).get("entityType")

    return {
        "periods": periods,
        "historicals": hist,
        "current": {
            "Diluted Shares Out": mm(shares),
            "Cash & Equivalents": mm(cash),
            "ST Investments": mm(sti),
            "Total Debt": mm(debt),
            "Preferred Equity": mm(preferred),
            "Minority Interest": mm(minority),
            "Equity Investments": mm(eq_inv_total) if eq_found else None,
            "Marketable Securities": mm(mkt_sec),
        },
        "balance_sheet_as_of": cash_end,
        "metadata": {
            "Company Name": facts.get("entityName"),
            "Sector": sic,
            "Currency": "USD",
            "Filing Status": category,
        },
    }


# --- Workbook write ----------------------------------------------------------
def write_model(model_path: Path, ticker: str, data: dict, price, price_as_of,
                notes: list[str]) -> None:
    wb = load_workbook(model_path)
    if "_CapIQ_Data" not in wb.sheetnames:
        raise SystemExit(
            "Model is missing _CapIQ_Data tab. "
            "Run `python -m shared.scaffold_template` and re-bootstrap."
        )
    ws = wb["_CapIQ_Data"]
    validate_field_labels(
        ws, capiq_layout.all_field_rows(), "_CapIQ_Data",
        "Fix: re-run `python -m shared.scaffold_template` and re-bootstrap, "
        "OR update shared/capiq_layout.py to match the workbook.",
    )

    meta_row = {label: r for r, label, _ in capiq_layout.METADATA}
    cur_row = {label: r for r, label, _ in capiq_layout.CURRENT_STATE}
    hist_row = {label: r for r, label in capiq_layout.HISTORICALS}
    col_cur = capiq_layout.COL_CURRENT
    hist_cols = (capiq_layout.COL_FY_M2, capiq_layout.COL_FY_M1,
                 capiq_layout.COL_FY)

    # NB: assign via .value everywhere — ws.cell(r, c, None) silently skips
    # the write (openpyxl only assigns non-None), leaving the template's
    # sample data in place instead of blanking the cell.
    for label, value in data["metadata"].items():
        ws.cell(meta_row[label], col_cur).value = value

    cur = data["current"]
    ws.cell(cur_row["Current Price"], col_cur).value = price
    for label in ("Diluted Shares Out", "Cash & Equivalents", "ST Investments",
                  "Total Debt", "Preferred Equity", "Minority Interest",
                  "Equity Investments", "Marketable Securities"):
        ws.cell(cur_row[label], col_cur).value = cur[label]

    # Market Cap / EV as hard values (matching fetch_capiq's verbatim-values
    # copy): extract scripts read with data_only=True, and openpyxl-written
    # formulas would have no cached values until the file is opened in Excel.
    shares = cur["Diluted Shares Out"]
    mkt_cap = price * shares if price is not None and shares is not None else None
    ws.cell(cur_row["Market Cap"], col_cur).value = mkt_cap
    ev = None
    if mkt_cap is not None:
        z = lambda v: v or 0.0
        ev = (mkt_cap - z(cur["Cash & Equivalents"]) - z(cur["ST Investments"])
              + z(cur["Total Debt"]) + z(cur["Preferred Equity"])
              + z(cur["Minority Interest"]) - z(cur["Equity Investments"])
              - z(cur["Marketable Securities"]))
    ws.cell(cur_row["Enterprise Value"], col_cur).value = ev

    for label, values in data["historicals"].items():
        for c, v in zip(hist_cols, values):
            ws.cell(hist_row[label], c).value = v

    # Stamps: ticker, fiscal-period ends under the Date row, run timestamp.
    ws.cell(capiq_layout.ROW_TICKER, 3, ticker)
    pad = 3 - len(data["periods"])
    for i, (_, end) in enumerate(data["periods"]):
        ws.cell(capiq_layout.ROW_DATE, hist_cols[pad + i], end.isoformat())
    if price_as_of:
        ws.cell(capiq_layout.ROW_DATE, col_cur, price_as_of)
    ts_cell = ws.cell(capiq_layout.ROW_FETCHER_DATE, 3, datetime.now())
    ts_cell.number_format = "mm/dd/yyyy hh:mm"

    wb.save(model_path)


# --- Main --------------------------------------------------------------------
def fetch(ticker: str, model_path_override: str | None = None,
          price_override: float | None = None, offline: bool = False) -> None:
    model_path = resolve_model_path(ticker, model_path_override)
    out_dir = model_path.parent
    fs = fs_ticker(ticker)
    notes: list[str] = []

    user_agent = os.environ.get("SEC_USER_AGENT")
    if not user_agent and not offline:
        raise SystemExit(
            "SEC_USER_AGENT is not set. Add a line like\n"
            "  SEC_USER_AGENT=Your Name you@example.com\n"
            "to the .env at the repo root (SEC requires an identifying UA)."
        )
    session = SecSession(user_agent or "offline")

    print(f"Writing XBRL values to: {model_path}")

    ticker_map = _fetch_cached(
        lambda: session.get_json(SEC_TICKER_MAP_URL),
        out_dir / "sec_company_tickers.json", offline, "ticker map")
    cik = None
    for entry in ticker_map.values():
        if str(entry.get("ticker", "")).upper() == ticker.upper():
            cik = int(entry["cik_str"])
            break
    if cik is None:
        raise SystemExit(f"Ticker {ticker} not found in SEC company_tickers.json "
                         f"(non-US filer or wrong symbol?)")
    print(f"  CIK: {cik}")

    facts = _fetch_cached(
        lambda: session.get_json(SEC_COMPANYFACTS_URL.format(cik=cik)),
        out_dir / f"{fs}_companyfacts.json", offline, "companyfacts")
    submissions = _fetch_cached(
        lambda: session.get_json(SEC_SUBMISSIONS_URL.format(cik=cik)),
        out_dir / f"{fs}_submissions.json", offline, "submissions",
        required=False)

    data = extract_financials(facts, submissions, notes)

    price = price_override
    price_as_of = None
    if price is None and not offline:
        try:
            quote = yahoo_last_close(ticker)
            if not _name_matches(data["metadata"]["Company Name"], quote["name"]):
                notes.append(
                    f"Yahoo resolved {ticker} to {quote['name']!r}, which does "
                    f"not look like {data['metadata']['Company Name']!r} — "
                    f"price left blank (recycled ticker?)")
            else:
                price = quote["price"]
                price_as_of = quote["as_of"]
        except Exception as e:  # price is best-effort; the model can wait
            notes.append(f"Yahoo price fetch failed ({e}) — pass --price to set "
                         f"manually")
    elif price is not None:
        price_as_of = date.today().isoformat()
        notes.append("price set via --price override")
    elif offline:
        notes.append("offline run — price blank unless --price given")

    write_model(model_path, ticker, data, price, price_as_of, notes)

    # --- Console summary ---
    h, cur = data["historicals"], data["current"]
    print(f"\nXBRL fetch complete: {ticker} "
          f"({data['metadata']['Company Name']})")
    print("  Fiscal years (cols C/D/E): " + " | ".join(
        f"{_fy_label(end)} ended {end}" for _, end in data["periods"]))
    print(f"  Balance sheet as of: {data['balance_sheet_as_of']}")

    def f(v, dec=1):
        return "blank" if v is None else f"{v:,.{dec}f}"

    print("\n  Sample values ($mm):")
    print(f"    Revenue (FY):     {f(h['Revenue'][2])}")
    print(f"    Adj EBITDA (FY):  {f(h['EBITDA'][2])}")
    print(f"    Cash:             {f(cur['Cash & Equivalents'])}")
    print(f"    Total Debt:       {f(cur['Total Debt'])}")
    print(f"    Shares out (mm):  {f(cur['Diluted Shares Out'], 2)}")
    print(f"    Price:            {f(price, 2)}"
          + (f"  (as of {price_as_of})" if price_as_of else ""))
    if notes:
        print("\n  Notes:")
        for n in notes:
            print(f"    - {n}")


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Fill _CapIQ_Data in a ticker's model from SEC XBRL + "
                    "Yahoo Finance (no CapIQ needed).")
    parser.add_argument("ticker", help="US-listed ticker (e.g. AAPL, LULU)")
    parser.add_argument("--price", type=float, default=None,
                        help="Manual current price (skips Yahoo).")
    parser.add_argument("--offline", action="store_true",
                        help="Use cached SEC downloads only; no network.")
    parser.add_argument("--model-path", default=None,
                        help="Override the model file. Default: the per-ticker "
                             "copy created by new_ticker.")
    args = parser.parse_args(argv)
    ticker = validate_ticker(args.ticker)
    fetch(ticker, model_path_override=args.model_path,
          price_override=args.price, offline=args.offline)


if __name__ == "__main__":
    main()
