"""Scaffold templates/capiq_fetcher.xlsx — standalone live-CapIQ workbook.

Idempotent: re-running overwrites the file. Runtime workflow:

    1. Open in Excel with the CapIQ plugin loaded.
    2. Set the ticker in B3 (named range `fetcher_ticker`).
    3. CapIQ resolves all formulas asynchronously.
    4. shared/fetch_capiq.py reads the values and pushes them into
       company_model.xlsx -> _CapIQ_Data tab.

Layout MUST mirror _CapIQ_Data on the main template. Both sources read
from shared.capiq_layout to keep them in lockstep.

Run:
    python -m shared.scaffold_capiq_fetcher
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName

from shared import capiq_layout


REPO_ROOT = Path(__file__).resolve().parent.parent
FETCHER_PATH = REPO_ROOT / "templates" / "capiq_fetcher.xlsx"
TICKER_REF = "fetcher_ticker"

# --- Styling (kept local; deliberately not importing scaffold_template's
#     styles to keep the fetcher script standalone-runnable) ---
BLUE = "0000FF"
WHITE = "FFFFFF"
YELLOW = "FFFF00"
HEADER_HEX = "1F4E78"
DOTTED_BLUE = Side(border_style="dotted", color=BLUE)

INPUT_FONT = Font(color=BLUE, name="Calibri", size=11)
INPUT_FILL = PatternFill("solid", fgColor=YELLOW)
INPUT_BORDER = Border(left=DOTTED_BLUE, right=DOTTED_BLUE, top=DOTTED_BLUE, bottom=DOTTED_BLUE)
HEADER_FONT = Font(color=WHITE, bold=True, name="Calibri", size=11)
HEADER_FILL = PatternFill("solid", fgColor=HEADER_HEX)
BANNER_FONT = Font(italic=True, color="808080", bold=True, name="Calibri", size=10)
LABEL_BOLD = Font(bold=True, name="Calibri", size=11)
FORMULA_FONT = Font(color="000000", name="Calibri", size=11)


def style_input(cell):
    cell.font = INPUT_FONT
    cell.fill = INPUT_FILL
    cell.border = INPUT_BORDER


def style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")


def style_formula(cell):
    cell.font = FORMULA_FONT


def build_fetcher(ws):
    # Banner / instructions
    ws.cell(capiq_layout.ROW_BANNER, 1,
            "CapIQ Fetcher — formulas only. Open in Excel with CapIQ plugin loaded to refresh."
            ).font = BANNER_FONT
    ws.cell(capiq_layout.ROW_LAST_FETCH, 1, "Run via:").font = LABEL_BOLD
    ws.cell(capiq_layout.ROW_LAST_FETCH, 2, "python -m shared.fetch_capiq <TICKER>")

    ws.cell(capiq_layout.ROW_TICKER, 1, "Ticker:").font = LABEL_BOLD
    ticker_cell = ws.cell(capiq_layout.ROW_TICKER, 2, "AAPL")
    style_input(ticker_cell)
    ticker_cell.comment = Comment(
        "Set this to the ticker you want to fetch. Workbook will refresh CapIQ "
        "formulas automatically when you change it (and when fetch_capiq.py "
        "drives it programmatically).",
        "scaffold_capiq_fetcher",
    )

    for j, h in enumerate(capiq_layout.COL_HEADERS):
        style_header(ws.cell(capiq_layout.ROW_COL_HEADERS, j + 1, h))

    # Section A: metadata (E column)
    for r, label, fmt in capiq_layout.METADATA:
        ws.cell(r, 1, label)
        formula = fmt.format(t=TICKER_REF)
        style_formula(ws.cell(r, 5, formula))

    # Section B: current (E column)
    for r, label, fmt in capiq_layout.CURRENT:
        ws.cell(r, 1, label)
        formula = fmt.format(t=TICKER_REF)
        style_formula(ws.cell(r, 5, formula))

    # Section C: historicals (B/C/D columns = FY-3/FY-2/FY-1)
    for r, label, func in capiq_layout.HISTORICAL:
        ws.cell(r, 1, label)
        if func is None:
            # Total OpEx (row 30) — sum SG&A + R&D
            for col_letter in capiq_layout.HIST_COLS:
                cell = ws[f"{col_letter}{r}"]
                cell.value = f"={col_letter}28+{col_letter}29"
                style_formula(cell)
        else:
            for col_letter, period in zip(capiq_layout.HIST_COLS, capiq_layout.HIST_PERIODS):
                cell = ws[f"{col_letter}{r}"]
                cell.value = f"={func}({TICKER_REF},{period},IQ_USD)"
                style_formula(cell)

    ws.column_dimensions["A"].width = 35
    for letter in ["B", "C", "D", "E"]:
        ws.column_dimensions[letter].width = 16


def build():
    wb = Workbook()
    wb.remove(wb.active)
    fetcher = wb.create_sheet("Fetcher")
    build_fetcher(fetcher)

    wb.defined_names[TICKER_REF] = DefinedName(
        name=TICKER_REF,
        attr_text=f"Fetcher!$B${capiq_layout.ROW_TICKER}",
    )

    FETCHER_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(FETCHER_PATH)
    return FETCHER_PATH


if __name__ == "__main__":
    out = build()
    print(f"Wrote {out}")
