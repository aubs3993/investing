"""Shared Excel styling constants, style helpers, and small fetch/extract utilities.

Single source of truth for the banker styling conventions (Arial 10pt,
blue-on-yellow inputs with dotted hair border, white-on-blue headers) used by
the scaffolders and the hardcoded-copy writers. Keep every workbook-producing
script importing from here so a styling tweak lands everywhere at once.
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# --- Universal styling (Arial 10pt across all generated workbooks) ---
ARIAL = "Arial"
ARIAL_SIZE = 10

BLUE = "0000FF"
WHITE = "FFFFFF"
YELLOW = "FFFF99"
HEADER_HEX = "0070C0"
HAIR = Side(border_style="hair")

INPUT_FONT = Font(color=BLUE, name=ARIAL, size=ARIAL_SIZE)
INPUT_FILL = PatternFill("solid", fgColor=YELLOW)
INPUT_BORDER = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
HEADER_FONT = Font(color=WHITE, bold=True, name=ARIAL, size=ARIAL_SIZE)
HEADER_FILL = PatternFill("solid", fgColor=HEADER_HEX)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
BANNER_FONT = Font(italic=True, color="808080", bold=True, name=ARIAL, size=ARIAL_SIZE)
LABEL_BOLD = Font(bold=True, name=ARIAL, size=ARIAL_SIZE)
LABEL = Font(name=ARIAL, size=ARIAL_SIZE)
FORMULA_FONT = Font(color="000000", name=ARIAL, size=ARIAL_SIZE)
TITLE_FONT = Font(bold=True, name=ARIAL, size=ARIAL_SIZE)


def style_input(cell):
    cell.font = INPUT_FONT
    cell.fill = INPUT_FILL
    cell.border = INPUT_BORDER


def style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN


def style_formula(cell):
    cell.font = FORMULA_FONT


def count_errors(values_2d) -> tuple[int, list[str]]:
    """Count Excel error strings ('#N/A', '#NAME?', ...) in a 2D value block.

    Ranges must be read with xlwings' options(err_to_str=True) for error
    cells to arrive as strings at all — the default converts them to None.
    """
    err_count = 0
    samples: list[str] = []
    for row in values_2d or []:
        for v in row:
            if isinstance(v, str) and v.startswith("#"):
                err_count += 1
                if len(samples) < 5:
                    samples.append(v)
    return err_count, samples


def format_money(v):
    if isinstance(v, (int, float)):
        return f"${v:,.0f}M" if abs(v) >= 1_000_000 else f"${v:,.2f}"
    return repr(v)


def validate_field_labels(ws, expected_rows, sheet_name, fix_hint):
    """Cross-check column-B field labels on an openpyxl sheet against a layout module.

    `expected_rows` is [(row, label), ...] — e.g. capiq_layout.all_field_rows().
    Mirrors _validate_layout_match in the fetch scripts so the read side aborts
    on layout drift instead of silently extracting the wrong cells.
    """
    misaligned = []
    for r, expected in expected_rows:
        actual = ws.cell(r, 2).value
        actual_l = str(actual).strip().lower() if actual is not None else ""
        if actual_l != expected.strip().lower():
            misaligned.append((r, expected, actual))
    if misaligned:
        lines = [f"Layout mismatch on {sheet_name} — extraction would read the wrong cells."]
        for r, expected, actual in misaligned:
            lines.append(f"  row {r}: expected {expected!r}, found {actual!r}")
        lines.append(fix_hint)
        raise SystemExit("\n".join(lines))
