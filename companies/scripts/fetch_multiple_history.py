"""Pull daily NTM and 2Y-forward EV/EBITDA multiple + growth history.

Usage:
    python -m companies.scripts.fetch_multiple_history <TICKER>
    python -m companies.scripts.fetch_multiple_history <TICKER> --end-date 2026-04-28
    python -m companies.scripts.fetch_multiple_history <TICKER> --lookback-years 3
    python -m companies.scripts.fetch_multiple_history <TICKER> --no-chart --headless

Workflow:
    1. Generate the business-day list (USFederalHolidayCalendar) from the
       lookback window ending on --end-date.
    2. Open templates/multiple_history_fetcher.xlsx, set ticker / end date /
       lookback inputs, write the date list into column B starting row 11.
    3. Force calc, wait for CapIQ async resolution, sleep 5s buffer.
    4. Read populated rows back as values; write a hardcoded copy to
       companies/output/<TICKER>/multiple_history_<TICKER>.xlsx.
    5. Build three dual-axis matplotlib charts unless --no-chart:
         - multiple_history_<TICKER>.png : Stock Price + NTM EV/EBITDA
         - multiple_growth_<TICKER>.png  : NTM EV/EBITDA + NTM Growth %
         - multiple_2y_<TICKER>.png      : 2Y Fwd EV/EBITDA + 2Y Fwd Growth (CAGR)
"""
from __future__ import annotations

import argparse
import re
import sys
import time
import traceback
from datetime import datetime, date as date_cls
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string
from openpyxl.workbook.defined_name import DefinedName

from shared import multiple_history_layout as layout
from shared.excel_session import AppPrefs, get_or_create_app, workbook_already_open

REPO_ROOT = Path(__file__).resolve().parents[2]
FETCHER_PATH = REPO_ROOT / "templates" / "multiple_history_fetcher.xlsx"
ASYNC_BUFFER_SECS = 5
TICKER_RE = re.compile(r"^[A-Z][A-Z0-9.\-:]{0,14}$")

# Styles for the hardcoded copy (mirror the fetcher scaffolder so the two
# look identical apart from values vs. formulas).
ARIAL = "Arial"
ARIAL_SIZE = 10
HAIR = Side(border_style="hair")
HEADER_FONT = Font(color="FFFFFF", bold=True, name=ARIAL, size=ARIAL_SIZE)
HEADER_FILL = PatternFill("solid", fgColor="0070C0")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
INPUT_FONT = Font(color="0000FF", name=ARIAL, size=ARIAL_SIZE)
INPUT_FILL = PatternFill("solid", fgColor="FFFF99")
INPUT_BORDER = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
LABEL_BOLD = Font(bold=True, name=ARIAL, size=ARIAL_SIZE)
LABEL = Font(name=ARIAL, size=ARIAL_SIZE)
TITLE_FONT = Font(bold=True, name=ARIAL, size=ARIAL_SIZE)
BANNER_FONT = Font(italic=True, color="808080", bold=True, name=ARIAL, size=ARIAL_SIZE)
VALUE_FONT = Font(color="000000", name=ARIAL, size=ARIAL_SIZE)


def _validate_ticker(raw: str) -> str:
    t = (raw or "").strip().upper()
    if not TICKER_RE.match(t):
        raise SystemExit(f"Invalid ticker: {raw!r}. Expected something like AAPL, BRK.B, 700:HK.")
    return t


def _parse_date(s: str) -> date_cls:
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        raise SystemExit(f"--end-date must be YYYY-MM-DD; got {s!r}")


def _generate_business_days(end: date_cls, lookback_years: int) -> list[date_cls]:
    try:
        import pandas as pd
        from pandas.tseries.holiday import USFederalHolidayCalendar
    except ImportError:
        raise SystemExit(
            "pandas is required. Install via `pip install -r requirements.txt`."
        )
    end_ts = pd.Timestamp(end)
    start_ts = end_ts - pd.DateOffset(years=lookback_years)
    holidays = USFederalHolidayCalendar().holidays(start=start_ts, end=end_ts)
    dates = pd.bdate_range(start=start_ts, end=end_ts, freq="C", holidays=holidays)
    return sorted([d.date() for d in dates], reverse=True)  # most recent first


def _check_capiq_loaded(sheet) -> None:
    """C11 holds an IFERROR-wrapped IQ_CLOSEPRICE call. After triggering calc,
    if the underlying CIQ name didn't resolve we'll see #NAME? before the
    IFERROR has a chance to mask anything (because IFERROR doesn't catch
    parse-time errors). Treat any string starting with '#' as a failure.
    """
    val = sheet.range((layout.ROW_DATA_START, 3)).value  # C11
    if isinstance(val, str) and val.strip().startswith("#NAME"):
        raise SystemExit(
            "CapIQ plugin not loaded (C11 returned #NAME?). Open Excel manually, "
            "sign in to the S&P Capital IQ plugin, then retry."
        )


def _count_errors(values_2d) -> tuple[int, list[str]]:
    err_count = 0
    samples: list[str] = []
    for row in values_2d or []:
        for v in row:
            if isinstance(v, str) and v.startswith("#"):
                err_count += 1
                if len(samples) < 5:
                    samples.append(v)
    return err_count, samples


def _last_data_col_idx() -> int:
    """Index of the last data column (currently X = 24)."""
    return column_index_from_string(layout.DATA_COLUMNS[-1][0])


def _style_header_cell(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN


def _build_hardcoded_copy(out_path: Path, ticker: str, end_date: date_cls,
                          lookback_years: int, dates: list[date_cls],
                          values_2d: list[list]) -> None:
    """Write a self-contained xlsx with the fetched values (no formulas)."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Fetcher")
    ws.sheet_view.showGridLines = False

    for letter, width in layout.COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    title = ws.cell(layout.ROW_TITLE, 2, f"{ticker} | NTM EV/EBITDA History")
    title.font = TITLE_FONT

    ws.cell(layout.ROW_BANNER, 2,
            "Hardcoded values from CapIQ. Generated by fetch_multiple_history.py."
            ).font = BANNER_FONT

    ws.cell(layout.ROW_RUN_VIA, 2, "Run via:").font = LABEL_BOLD
    ws.cell(layout.ROW_RUN_VIA, 3,
            "python -m companies.scripts.fetch_multiple_history <TICKER>").font = LABEL

    ws.cell(layout.ROW_TICKER, 2, "Ticker:").font = LABEL_BOLD
    tcell = ws.cell(layout.ROW_TICKER, 3, ticker)
    tcell.font = INPUT_FONT
    tcell.fill = INPUT_FILL
    tcell.border = INPUT_BORDER

    ws.cell(layout.ROW_END_DATE, 2, "End Date:").font = LABEL_BOLD
    ecell = ws.cell(layout.ROW_END_DATE, 3, end_date)
    ecell.font = INPUT_FONT
    ecell.fill = INPUT_FILL
    ecell.border = INPUT_BORDER
    ecell.number_format = "mm/dd/yyyy"

    ws.cell(layout.ROW_LOOKBACK_YRS, 2, "Lookback Years:").font = LABEL_BOLD
    lcell = ws.cell(layout.ROW_LOOKBACK_YRS, 3, lookback_years)
    lcell.font = INPUT_FONT
    lcell.fill = INPUT_FILL
    lcell.border = INPUT_BORDER
    lcell.number_format = "0"

    ws.cell(layout.ROW_GENERATED, 2, "Generated:").font = LABEL_BOLD
    gcell = ws.cell(layout.ROW_GENERATED, 3, datetime.now())
    gcell.font = LABEL
    gcell.number_format = "mm/dd/yyyy hh:mm"

    for letter, header, _role in layout.DATA_COLUMNS:
        col_idx = column_index_from_string(letter)
        _style_header_cell(ws.cell(layout.ROW_COL_HEADERS, col_idx, header))

    last_col_idx = _last_data_col_idx()
    n = len(dates)
    for i in range(n):
        r = layout.ROW_DATA_START + i
        # Date in column B.
        dcell = ws.cell(r, 2, dates[i])
        dcell.font = LABEL
        dcell.number_format = layout.COLUMN_NUMBER_FORMATS["B"]
        # Remaining columns from values_2d. Each row of values_2d corresponds
        # to a fetcher row; values_2d[i] is the row for dates[i].
        row_vals = values_2d[i] if i < len(values_2d) else []
        # values_2d came from columns B..O — index 0 is the date (already
        # written above), so iterate from index 1.
        for letter, _label, _role in layout.DATA_COLUMNS[1:]:
            col_idx = column_index_from_string(letter)
            v = row_vals[col_idx - 2] if (col_idx - 2) < len(row_vals) else None
            cell = ws.cell(r, col_idx, v)
            cell.font = VALUE_FONT
            cell.number_format = layout.COLUMN_NUMBER_FORMATS[letter]

    ws.freeze_panes = "A11"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


COLOR_PRICE = "#1f4e78"
COLOR_MULT = "#c00000"
COLOR_GROWTH = "#2e7d32"
MIN_CHART_ROWS = 50


def _import_chart_libs():
    try:
        import pandas as pd
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
    except ImportError:
        raise SystemExit(
            "matplotlib + pandas required for charting. "
            "Install via `pip install -r requirements.txt`, or pass --no-chart."
        )
    return pd, plt, mdates


def _read_chart_df(out_xlsx: Path):
    """Header row 10 (1-indexed) = pandas header=9."""
    pd, _, _ = _import_chart_libs()
    df = pd.read_excel(out_xlsx, sheet_name="Fetcher", header=9)
    df = df.rename(columns=lambda c: str(c).strip())
    df = df.dropna(subset=["Date"])
    return df.sort_values("Date").reset_index(drop=True)


def _filter_multiple(df, mult_col):
    """Keep rows with multiples strictly between 0x and 100x."""
    s = df[mult_col]
    return df[(s > 0) & (s < 100)].copy()


def _filter_growth(df, growth_col, low=-1.0, high=5.0):
    s = df[growth_col]
    return df[(s >= low) & (s <= high)].copy()


def _format_y_axis(ax, color, label):
    ax.set_ylabel(label, color=color)
    ax.tick_params(axis="y", labelcolor=color)


def _format_x_dates(ax, mdates):
    ax.xaxis.set_major_locator(mdates.YearLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))


def _combine_legends(ax1, ax2):
    l1, lab1 = ax1.get_legend_handles_labels()
    l2, lab2 = ax2.get_legend_handles_labels()
    ax1.legend(l1 + l2, lab1 + lab2, loc="upper left", fontsize=9)


def _chart_price_and_ntm_multiple(df, ticker, out_path):
    pd, plt, mdates = _import_chart_libs()
    sub = _filter_multiple(df, "NTM EV/EBITDA").dropna(subset=["Stock Price"])
    if len(sub) < MIN_CHART_ROWS:
        return {"skipped": True, "rows": len(sub), "reason": "insufficient rows"}

    plt.rcParams.update({"font.family": "Arial", "font.size": 10})
    fig, ax1 = plt.subplots(figsize=(12, 6))

    ax1.plot(sub["Date"], sub["Stock Price"], color=COLOR_PRICE, linewidth=1.2, label="Stock Price")
    ax1.set_xlabel("Date")
    _format_y_axis(ax1, COLOR_PRICE, "Stock Price ($)")
    ax1.grid(True, alpha=0.3)

    ax2 = ax1.twinx()
    ax2.plot(sub["Date"], sub["NTM EV/EBITDA"], color=COLOR_MULT, linewidth=1.2,
             label="NTM EV/EBITDA")
    median_mult = float(sub["NTM EV/EBITDA"].median())
    ax2.axhline(median_mult, color=COLOR_MULT, linestyle="--", alpha=0.4,
                linewidth=0.8, label=f"Median: {median_mult:.1f}x")
    _format_y_axis(ax2, COLOR_MULT, "NTM EV/EBITDA Multiple")

    plt.title(f"{ticker} — Stock Price and NTM EV/EBITDA Multiple",
              fontsize=12, fontweight="bold")
    _format_x_dates(ax1, mdates)
    fig.autofmt_xdate()
    _combine_legends(ax1, ax2)

    plt.tight_layout()
    plt.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close()
    return {"skipped": False, "rows": len(sub), "median_multiple": median_mult}


def _chart_ntm_multiple_and_growth(df, ticker, out_path):
    pd, plt, mdates = _import_chart_libs()
    sub = _filter_multiple(df, "NTM EV/EBITDA")
    sub = _filter_growth(sub, "NTM Growth %")
    if len(sub) < MIN_CHART_ROWS:
        return {"skipped": True, "rows": len(sub), "reason": "insufficient rows"}

    plt.rcParams.update({"font.family": "Arial", "font.size": 10})
    fig, ax1 = plt.subplots(figsize=(12, 6))

    ax1.plot(sub["Date"], sub["NTM EV/EBITDA"], color=COLOR_MULT, linewidth=1.2,
             label="NTM EV/EBITDA")
    median_mult = float(sub["NTM EV/EBITDA"].median())
    ax1.axhline(median_mult, color=COLOR_MULT, linestyle="--", alpha=0.4,
                linewidth=0.8, label=f"Median multiple: {median_mult:.1f}x")
    _format_y_axis(ax1, COLOR_MULT, "NTM EV/EBITDA Multiple")
    ax1.grid(True, alpha=0.3)

    ax2 = ax1.twinx()
    ax2.plot(sub["Date"], sub["NTM Growth %"], color=COLOR_GROWTH, linewidth=1.2,
             label="NTM Growth %")
    median_growth = float(sub["NTM Growth %"].median())
    ax2.axhline(median_growth, color=COLOR_GROWTH, linestyle="--", alpha=0.4,
                linewidth=0.8, label=f"Median growth: {median_growth:.1%}")
    ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:.0%}"))
    _format_y_axis(ax2, COLOR_GROWTH, "NTM EBITDA Growth %")

    plt.title(f"{ticker} — NTM Multiple vs. NTM Growth Expectations",
              fontsize=12, fontweight="bold")
    _format_x_dates(ax1, mdates)
    fig.autofmt_xdate()
    _combine_legends(ax1, ax2)

    plt.tight_layout()
    plt.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close()
    return {"skipped": False, "rows": len(sub),
            "median_multiple": median_mult, "median_growth": median_growth}


def _chart_2y_forward(df, ticker, out_path):
    pd, plt, mdates = _import_chart_libs()
    sub = _filter_multiple(df, "2Y Fwd EV/EBITDA")
    sub = _filter_growth(sub, "2Y Fwd Growth % (CAGR)")
    if len(sub) < MIN_CHART_ROWS:
        return {"skipped": True, "rows": len(sub),
                "reason": "insufficient rows (likely sparse IQ_CY+3 broker coverage)"}

    plt.rcParams.update({"font.family": "Arial", "font.size": 10})
    fig, ax1 = plt.subplots(figsize=(12, 6))

    ax1.plot(sub["Date"], sub["2Y Fwd EV/EBITDA"], color=COLOR_MULT, linewidth=1.2,
             label="2Y Fwd EV/EBITDA")
    median_mult = float(sub["2Y Fwd EV/EBITDA"].median())
    ax1.axhline(median_mult, color=COLOR_MULT, linestyle="--", alpha=0.4,
                linewidth=0.8, label=f"Median multiple: {median_mult:.1f}x")
    _format_y_axis(ax1, COLOR_MULT, "2Y Forward EV/EBITDA Multiple")
    ax1.grid(True, alpha=0.3)

    ax2 = ax1.twinx()
    ax2.plot(sub["Date"], sub["2Y Fwd Growth % (CAGR)"], color=COLOR_GROWTH,
             linewidth=1.2, label="2Y Fwd Growth % (CAGR)")
    median_growth = float(sub["2Y Fwd Growth % (CAGR)"].median())
    ax2.axhline(median_growth, color=COLOR_GROWTH, linestyle="--", alpha=0.4,
                linewidth=0.8, label=f"Median growth: {median_growth:.1%}")
    ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:.0%}"))
    _format_y_axis(ax2, COLOR_GROWTH, "2Y Forward EBITDA Growth % (annualized)")

    plt.title(f"{ticker} — 2Y Forward Multiple vs. 2Y Forward Growth (CAGR)",
              fontsize=12, fontweight="bold")
    _format_x_dates(ax1, mdates)
    fig.autofmt_xdate()
    _combine_legends(ax1, ax2)

    plt.tight_layout()
    plt.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close()
    return {"skipped": False, "rows": len(sub),
            "median_multiple": median_mult, "median_growth": median_growth}


def _generate_charts(out_xlsx: Path, out_dir: Path, ticker: str) -> dict:
    """Build all three charts. Returns a dict keyed 'price_ntm', 'ntm_growth', '2y_fwd'."""
    df = _read_chart_df(out_xlsx)
    paths = {
        "price_ntm": out_dir / f"multiple_history_{ticker}.png",
        "ntm_growth": out_dir / f"multiple_growth_{ticker}.png",
        "2y_fwd": out_dir / f"multiple_2y_{ticker}.png",
    }
    results = {
        "price_ntm": _chart_price_and_ntm_multiple(df, ticker, paths["price_ntm"]),
        "ntm_growth": _chart_ntm_multiple_and_growth(df, ticker, paths["ntm_growth"]),
        "2y_fwd": _chart_2y_forward(df, ticker, paths["2y_fwd"]),
    }
    for key, info in results.items():
        info["path"] = paths[key]
    return results


def fetch(ticker: str, end_date: date_cls, lookback_years: int,
          no_chart: bool = False, headless: bool = False) -> None:
    if not FETCHER_PATH.exists():
        raise SystemExit(
            f"Missing {FETCHER_PATH}. Run "
            f"`python -m shared.scaffold_multiple_history_fetcher` first."
        )

    dates = _generate_business_days(end_date, lookback_years)
    if len(dates) > layout.data_capacity():
        raise SystemExit(
            f"{len(dates)} business days requested but the fetcher template "
            f"has capacity for {layout.data_capacity()}. Increase ROW_DATA_END "
            f"in shared/multiple_history_layout.py and rerun the scaffolder."
        )
    if not dates:
        raise SystemExit("No business days in the requested window.")

    out_dir = REPO_ROOT / "companies" / "output" / ticker
    out_xlsx = out_dir / f"multiple_history_{ticker}.xlsx"
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"Multiple history fetch: {ticker}")
    print(f"  Date range: {dates[-1]} to {dates[0]}  ({len(dates)} business days)")
    print(f"  Output dir: {out_dir}")

    app, owns_app = get_or_create_app(headless=headless)
    if not owns_app and workbook_already_open(app, FETCHER_PATH):
        raise SystemExit(
            "multiple_history_fetcher.xlsx is already open in your Excel. "
            "Close it before running this script."
        )

    fetcher_wb = None
    last_col_idx = _last_data_col_idx()
    try:
        prefs = AppPrefs(app)
        prefs.__enter__()

        fetcher_wb = app.books.open(str(FETCHER_PATH), update_links=False)
        fetcher_sheet = fetcher_wb.sheets["Fetcher"]

        # Set inputs via named ranges (fall back to direct cells if names
        # are missing for any reason).
        try:
            fetcher_wb.names[layout.NAME_TICKER].refers_to_range.value = ticker
        except Exception:
            fetcher_sheet.range((layout.ROW_TICKER, 3)).value = ticker
        try:
            fetcher_wb.names[layout.NAME_END_DATE].refers_to_range.value = end_date
        except Exception:
            fetcher_sheet.range((layout.ROW_END_DATE, 3)).value = end_date
        try:
            fetcher_wb.names[layout.NAME_LOOKBACK_YRS].refers_to_range.value = lookback_years
        except Exception:
            fetcher_sheet.range((layout.ROW_LOOKBACK_YRS, 3)).value = lookback_years

        # Bulk-write dates into column B starting at ROW_DATA_START.
        date_block = [[d] for d in dates]
        fetcher_sheet.range(
            (layout.ROW_DATA_START, 2),
            (layout.ROW_DATA_START + len(dates) - 1, 2),
        ).value = date_block

        # Clear leftover dates beyond len(dates) up to ROW_DATA_END.
        if layout.ROW_DATA_START + len(dates) <= layout.ROW_DATA_END:
            blanks = [[None]] * (layout.ROW_DATA_END - (layout.ROW_DATA_START + len(dates)) + 1)
            fetcher_sheet.range(
                (layout.ROW_DATA_START + len(dates), 2),
                (layout.ROW_DATA_END, 2),
            ).value = blanks

        # Trigger calc + async resolution.
        app.calculate()
        try:
            app.api.CalculateUntilAsyncQueriesDone()
        except Exception:
            time.sleep(8)
        time.sleep(ASYNC_BUFFER_SECS)

        _check_capiq_loaded(fetcher_sheet)

        # Read back populated range B..X for len(dates) rows.
        last_row = layout.ROW_DATA_START + len(dates) - 1
        values = fetcher_sheet.range(
            (layout.ROW_DATA_START, 2),
            (last_row, last_col_idx),
        ).value
        if values is None:
            values = []
        # xlwings returns a flat list when there's only one row — normalize.
        if values and not isinstance(values[0], list):
            values = [values]

        err_count, samples = _count_errors(values)
        print(f"  CapIQ async resolved.  Errors: {err_count}"
              + (f"  e.g. {samples}" if samples else ""))

        # Build the standalone hardcoded copy.
        _build_hardcoded_copy(
            out_xlsx, ticker, end_date, lookback_years, dates, values,
        )
        print(f"  Wrote: {out_xlsx}")

    except SystemExit:
        raise
    except Exception:
        print("Unhandled error during fetch:", file=sys.stderr)
        traceback.print_exc()
        raise SystemExit(1)
    finally:
        try:
            if fetcher_wb is not None:
                fetcher_wb.close()  # don't save — keep fetcher template clean
        except Exception:
            pass
        try:
            prefs.__exit__(None, None, None)
        except Exception:
            pass
        if owns_app:
            try:
                app.quit()
            except Exception:
                pass

    print()
    print(f"Multiple history complete: {ticker}")
    print(f"  Date range: {dates[-1]} to {dates[0]}")
    print(f"  Business days fetched: {len(dates)}")

    if no_chart:
        print()
        print("Files written:")
        print(f"  {out_xlsx}")
        return

    chart_results = _generate_charts(out_xlsx, out_dir, ticker)
    print()
    c1 = chart_results["price_ntm"]
    c2 = chart_results["ntm_growth"]
    c3 = chart_results["2y_fwd"]

    if not c1["skipped"]:
        print(f"  Median NTM EV/EBITDA: {c1['median_multiple']:.1f}x  (rows: {c1['rows']})")
    else:
        print(f"  Chart 1 skipped — {c1.get('reason', 'no data')} ({c1['rows']} rows).")

    if not c2["skipped"]:
        print(f"  Median NTM Growth: {c2['median_growth']:.1%}       (rows: {c2['rows']})")
    else:
        print(f"  Chart 2 skipped — {c2.get('reason', 'no data')} ({c2['rows']} rows).")

    if not c3["skipped"]:
        print(f"  Median 2Y Fwd EV/EBITDA: {c3['median_multiple']:.1f}x (rows: {c3['rows']})")
        print(f"  Median 2Y Fwd Growth (CAGR): {c3['median_growth']:.1%} (rows: {c3['rows']})")
    else:
        print(f"  Chart 3 skipped — {c3.get('reason', 'no data')} ({c3['rows']} rows).")

    print()
    print("Files written:")
    print(f"  {out_xlsx}")
    for key in ("price_ntm", "ntm_growth", "2y_fwd"):
        info = chart_results[key]
        if not info["skipped"]:
            print(f"  {info['path']}")


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Pull daily NTM and 2Y-forward EV/EBITDA + growth history.",
    )
    parser.add_argument("ticker", help="Ticker (e.g. AAPL, BRK.B, 700:HK)")
    parser.add_argument("--end-date", default=None,
                        help="End of lookback window (YYYY-MM-DD). Defaults to today.")
    parser.add_argument("--lookback-years", type=int, default=5,
                        help="Number of years to look back. Default 5.")
    parser.add_argument("--no-chart", action="store_true",
                        help="Skip chart generation; write only the hardcoded xlsx.")
    parser.add_argument("--headless", action="store_true",
                        help="Run Excel hidden. Default is visible for CapIQ debugging.")
    args = parser.parse_args(argv)

    ticker = _validate_ticker(args.ticker)
    end_date = _parse_date(args.end_date) if args.end_date else date_cls.today()
    if args.lookback_years <= 0:
        raise SystemExit("--lookback-years must be positive.")

    fetch(ticker, end_date, args.lookback_years,
          no_chart=args.no_chart, headless=args.headless)


if __name__ == "__main__":
    main()
