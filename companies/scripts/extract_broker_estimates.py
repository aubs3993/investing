"""Extract broker estimates brief from a ticker's _Broker_Data tab.

Usage:
    python -m companies.scripts.extract_broker_estimates <TICKER>
    python -m companies.scripts.extract_broker_estimates <TICKER> --format markdown
    python -m companies.scripts.extract_broker_estimates <TICKER> --output PATH

Implied drivers (revenue growth, margins, capex %) are computed in the same
units as the model's drivers, so they can be compared directly.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from shared import broker_layout, capiq_layout
from shared.excel_helpers import validate_field_labels
from shared.tickers import fs_ticker

REPO_ROOT = Path(__file__).resolve().parents[2]

# Cell coordinates sourced from the layout modules (the source of truth for
# the hidden data tabs). Single-value cells (last fetch, FY years, sentiment)
# live in column C, same as the FY1 means.
PNL_ROW = {label: r for r, label, *_ in broker_layout.PNL}
SENT_ROW = {label: r for r, label, _ in broker_layout.SENTIMENT}
COL_FY1_MEAN, COL_FY2_MEAN, COL_FY3_MEAN = (
    column_index_from_string(c) for c in broker_layout.MEAN_COLS
)
COL_FY1_HIGH = column_index_from_string(broker_layout.HIGH_COL)
COL_FY1_LOW = column_index_from_string(broker_layout.LOW_COL)
COL_FY1_COUNT = column_index_from_string(broker_layout.COUNT_COL)
COL_VALUE = COL_FY1_MEAN

CAPIQ_HIST_ROW = {label: r for r, label in capiq_layout.HISTORICALS}
CAPIQ_CUR_ROW = {label: r for r, label, _ in capiq_layout.CURRENT_STATE}


def _model_path_for(ticker: str) -> Path:
    fs = fs_ticker(ticker)
    return REPO_ROOT / "companies" / "output" / fs / f"{fs}_model.xlsx"


def _safe_div(num, den):
    try:
        if den in (None, 0) or num is None:
            return None
        return num / den
    except (TypeError, ZeroDivisionError):
        return None


def _safe_sub(a, b):
    if a is None or b is None:
        return None
    return a - b


def _safe_growth(prev, curr):
    try:
        if prev in (None, 0) or curr is None:
            return None
        return curr / prev - 1
    except TypeError:  # non-numeric, e.g. '#N/A' error strings on the data tab
        return None


def _get(ws, row, col):
    return ws.cell(row, col).value


def _recommendation_label(score):
    if not isinstance(score, (int, float)):
        return None
    if score < 1.5:
        return "Strong Buy"
    if score < 2.5:
        return "Buy / Outperform"
    if score < 3.5:
        return "Hold"
    if score < 4.5:
        return "Underperform"
    return "Sell"


def extract(ticker: str) -> dict:
    model = _model_path_for(ticker)
    if not model.exists():
        raise SystemExit(
            f"Missing {model}. Bootstrap with `python -m companies.scripts.new_ticker {ticker}` "
            f"and then run `python -m shared.fetch_broker_estimates {ticker}`."
        )
    wb = load_workbook(model, data_only=True)
    if "_Broker_Data" not in wb.sheetnames:
        raise SystemExit("Model is missing _Broker_Data tab. Regenerate via scaffold_template.")
    if "_CapIQ_Data" not in wb.sheetnames:
        raise SystemExit("Model is missing _CapIQ_Data tab. Regenerate via scaffold_template.")
    bk = wb["_Broker_Data"]
    cap = wb["_CapIQ_Data"]

    # Abort on layout drift instead of silently reading the wrong cells.
    layout_hint = ("Fix: re-run `python -m shared.scaffold_template` and re-fetch, "
                   "OR update the layout module to match the workbook.")
    validate_field_labels(bk, broker_layout.all_field_rows(), "_Broker_Data", layout_hint)
    validate_field_labels(cap, capiq_layout.all_field_rows(), "_CapIQ_Data", layout_hint)

    fy_cal = {
        "FY1": _get(bk, broker_layout.ROW_FY1_YEAR, COL_VALUE),
        "FY2": _get(bk, broker_layout.ROW_FY2_YEAR, COL_VALUE),
        "FY3": _get(bk, broker_layout.ROW_FY3_YEAR, COL_VALUE),
    }

    # P&L grid (broker_layout.PNL rows): C/D/E = FY1/FY2/FY3 mean, F/G = FY1 high/low, H = FY1 #est
    def metric(row):
        fy1_mean = _get(bk, row, COL_FY1_MEAN)
        fy2_mean = _get(bk, row, COL_FY2_MEAN)
        fy3_mean = _get(bk, row, COL_FY3_MEAN)
        fy1_high = _get(bk, row, COL_FY1_HIGH)
        fy1_low = _get(bk, row, COL_FY1_LOW)
        fy1_n = _get(bk, row, COL_FY1_COUNT)
        dispersion = None
        if isinstance(fy1_high, (int, float)) and isinstance(fy1_low, (int, float)) and isinstance(fy1_mean, (int, float)) and fy1_mean:
            dispersion = (fy1_high - fy1_low) / fy1_mean
        return {
            "FY1": {
                "mean": fy1_mean, "high": fy1_high, "low": fy1_low,
                "n_estimates": fy1_n, "dispersion": dispersion,
            },
            "FY2": {"mean": fy2_mean},
            "FY3": {"mean": fy3_mean},
        }

    revenue = metric(PNL_ROW["Revenue"])
    gross_profit = metric(PNL_ROW["Gross Profit"])
    ebitda = metric(PNL_ROW["EBITDA"])
    ebit = metric(PNL_ROW["EBIT"])
    net_income = metric(PNL_ROW["Net Income"])
    eps = metric(PNL_ROW["EPS (Diluted)"])
    cfo = metric(PNL_ROW["CFO"])
    capex = metric(PNL_ROW["CapEx"])

    # Most recent completed-FY revenue from the CapIQ tab is the base for FY1
    # implied growth — the same base broker_layout.IMPLIED uses in the model.
    rev_fy = _get(cap, CAPIQ_HIST_ROW["Revenue"], capiq_layout.COL_FY)
    revenue["FY1"]["implied_growth"] = _safe_growth(rev_fy, revenue["FY1"]["mean"])
    revenue["FY2"]["implied_growth_yoy"] = _safe_growth(revenue["FY1"]["mean"], revenue["FY2"]["mean"])
    revenue["FY3"]["implied_growth_yoy"] = _safe_growth(revenue["FY2"]["mean"], revenue["FY3"]["mean"])

    implied_drivers = {
        "revenue_growth_FY1_FY3": [
            revenue["FY1"].get("implied_growth"),
            revenue["FY2"].get("implied_growth_yoy"),
            revenue["FY3"].get("implied_growth_yoy"),
        ],
        "gross_margin_FY1":   _safe_div(gross_profit["FY1"]["mean"], revenue["FY1"]["mean"]),
        "ebitda_margin_FY1":  _safe_div(ebitda["FY1"]["mean"], revenue["FY1"]["mean"]),
        "ebit_margin_FY1":    _safe_div(ebit["FY1"]["mean"], revenue["FY1"]["mean"]),
        "capex_pct_revenue_FY1": _safe_div(capex["FY1"]["mean"], revenue["FY1"]["mean"]),
    }

    # Sentiment block (broker_layout.SENTIMENT rows, values in column C)
    n_analysts = _get(bk, SENT_ROW["Number of Analysts Covering"], COL_VALUE)
    avg_target = _get(bk, SENT_ROW["Average Price Target"], COL_VALUE)
    implied_upside = _get(bk, SENT_ROW["Implied Upside %"], COL_VALUE)
    avg_rec = _get(bk, SENT_ROW["Average Recommendation (1=Buy,5=Sell)"], COL_VALUE)
    rec_dist = _get(bk, SENT_ROW["Recommendation Distribution"], COL_VALUE)

    current_price = _get(cap, CAPIQ_CUR_ROW["Current Price"], capiq_layout.COL_CURRENT)
    if implied_upside in (None, "") and isinstance(avg_target, (int, float)) and isinstance(current_price, (int, float)) and current_price:
        implied_upside = avg_target / current_price - 1

    return {
        "ticker": ticker,
        "fetch_timestamp": _get(bk, broker_layout.ROW_LAST_FETCH, COL_VALUE),
        "broker_estimates": {
            "fy_calendar": fy_cal,
            "revenue": revenue,
            "gross_profit": gross_profit,
            "ebitda": ebitda,
            "ebit": ebit,
            "net_income": net_income,
            "eps_diluted": eps,
            "cfo": cfo,
            "capex": capex,
            "implied_drivers": implied_drivers,
        },
        "analyst_sentiment": {
            "num_analysts": n_analysts,
            "avg_price_target": avg_target,
            "current_price": current_price,
            "implied_upside": implied_upside,
            "avg_recommendation": avg_rec,
            "recommendation_label": _recommendation_label(avg_rec),
            "recommendation_distribution": rec_dist,
        },
    }


def _fmt_pct(v): return f"{v*100:.1f}%" if isinstance(v, (int, float)) else "—"
def _fmt_num(v): return f"{v:,.0f}" if isinstance(v, (int, float)) else "—"
def _fmt_money(v): return f"${v:,.2f}" if isinstance(v, (int, float)) else "—"


def to_markdown(data: dict) -> str:
    be = data["broker_estimates"]
    sent = data["analyst_sentiment"]
    fy = be["fy_calendar"]
    lines = [
        f"# {data['ticker']} — Broker Estimates Brief",
        "",
        f"**Fetched:** {data.get('fetch_timestamp')}",
        f"**Fiscal year mapping:** FY1={fy.get('FY1')} | FY2={fy.get('FY2')} | FY3={fy.get('FY3')}",
        "",
        "## Consensus P&L",
        "",
        "| Metric | FY1 mean | FY2 mean | FY3 mean | FY1 high | FY1 low | FY1 #est | FY1 dispersion |",
        "|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for label, key in [
        ("Revenue", "revenue"), ("Gross Profit", "gross_profit"),
        ("EBITDA", "ebitda"), ("EBIT", "ebit"),
        ("Net Income", "net_income"), ("EPS Diluted", "eps_diluted"),
        ("CFO", "cfo"), ("CapEx", "capex"),
    ]:
        m = be[key]
        lines.append(
            f"| {label} | {_fmt_num(m['FY1']['mean'])} | {_fmt_num(m['FY2']['mean'])} | {_fmt_num(m['FY3']['mean'])} | "
            f"{_fmt_num(m['FY1'].get('high'))} | {_fmt_num(m['FY1'].get('low'))} | "
            f"{_fmt_num(m['FY1'].get('n_estimates'))} | {_fmt_pct(m['FY1'].get('dispersion'))} |"
        )

    impl = be["implied_drivers"]
    lines += [
        "",
        "## Implied drivers (consensus, in model units)",
        "",
        f"- Revenue growth FY1 → FY3: "
        f"{_fmt_pct(impl['revenue_growth_FY1_FY3'][0])}, "
        f"{_fmt_pct(impl['revenue_growth_FY1_FY3'][1])}, "
        f"{_fmt_pct(impl['revenue_growth_FY1_FY3'][2])}",
        f"- Gross margin FY1: {_fmt_pct(impl['gross_margin_FY1'])}",
        f"- EBITDA margin FY1: {_fmt_pct(impl['ebitda_margin_FY1'])}",
        f"- EBIT margin FY1: {_fmt_pct(impl['ebit_margin_FY1'])}",
        f"- CapEx % of revenue FY1: {_fmt_pct(impl['capex_pct_revenue_FY1'])}",
        "",
        "## Analyst sentiment",
        "",
        f"- Analysts covering: {_fmt_num(sent.get('num_analysts'))}",
        f"- Average price target: {_fmt_money(sent.get('avg_price_target'))}",
        f"- Current price: {_fmt_money(sent.get('current_price'))}",
        f"- Implied upside: {_fmt_pct(sent.get('implied_upside'))}",
        f"- Avg recommendation: {sent.get('avg_recommendation')} ({sent.get('recommendation_label')})",
    ]
    return "\n".join(lines)


def main(argv=None):
    parser = argparse.ArgumentParser(description="Extract broker estimates brief from _Broker_Data.")
    parser.add_argument("ticker")
    parser.add_argument("--format", choices=["json", "markdown"], default="json")
    parser.add_argument("--output", default=None)
    args = parser.parse_args(argv)
    ticker = args.ticker.strip().upper()
    data = extract(ticker)
    if args.format == "json":
        out = json.dumps(data, indent=2, default=str)
    else:
        out = to_markdown(data)
    if args.output:
        Path(args.output).write_text(out, encoding="utf-8")
        print(f"Wrote {args.output}")
    else:
        sys.stdout.write(out + "\n")


if __name__ == "__main__":
    main()
