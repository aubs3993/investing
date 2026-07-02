"""Extract historicals brief from a ticker's _CapIQ_Data tab.

Usage:
    python -m companies.scripts.extract_historicals <TICKER>
    python -m companies.scripts.extract_historicals <TICKER> --format markdown
    python -m companies.scripts.extract_historicals <TICKER> --output PATH
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from shared import capiq_layout
from shared.excel_helpers import validate_field_labels
from shared.tickers import fs_ticker

REPO_ROOT = Path(__file__).resolve().parents[2]

# Row positions sourced from shared/capiq_layout.py (the source of truth for
# the _CapIQ_Data tab).
HIST_ROW = {label: r for r, label in capiq_layout.HISTORICALS}
CUR_ROW = {label: r for r, label, _ in capiq_layout.CURRENT_STATE}
META_ROW = {label: r for r, label, _ in capiq_layout.METADATA}


def _model_path_for(ticker: str) -> Path:
    fs = fs_ticker(ticker)
    return REPO_ROOT / "companies" / "output" / fs / f"{fs}_model.xlsx"


def _safe_div(num, den):
    try:
        if den in (None, 0) or num is None:
            return None
        return num / den
    except (TypeError, ZeroDivisionError):
        return None


def _cagr(start, end, years):
    try:
        if start is None or end is None or start <= 0 or end <= 0 or years <= 0:
            return None
        return (end / start) ** (1 / years) - 1
    except TypeError:  # non-numeric, e.g. '#N/A' error strings on the data tab
        return None


def _yoy(prev, curr):
    try:
        if prev in (None, 0) or curr is None:
            return None
        return curr / prev - 1
    except TypeError:  # non-numeric, e.g. '#N/A' error strings on the data tab
        return None


def _avg(vals):
    clean = [v for v in vals if isinstance(v, (int, float))]
    if not clean:
        return None
    return sum(clean) / len(clean)


def _get(ws, row, col):
    return ws.cell(row, col).value


def extract(ticker: str) -> dict:
    model = _model_path_for(ticker)
    if not model.exists():
        raise SystemExit(
            f"Missing {model}. Bootstrap with `python -m companies.scripts.new_ticker {ticker}` "
            f"and then run `python -m shared.fetch_capiq {ticker}`."
        )
    wb = load_workbook(model, data_only=True)
    if "_CapIQ_Data" not in wb.sheetnames:
        raise SystemExit("Model is missing _CapIQ_Data tab. Regenerate via scaffold_template.")
    cap = wb["_CapIQ_Data"]

    # Abort on layout drift instead of silently reading the wrong cells.
    validate_field_labels(
        cap, capiq_layout.all_field_rows(), "_CapIQ_Data",
        "Fix: re-run `python -m shared.scaffold_template` and re-fetch, "
        "OR update shared/capiq_layout.py to match the workbook.",
    )

    # Historical rows on _CapIQ_Data: cols C/D/E = FY-2, FY-1, FY (most recent
    # completed year). Two-period span (FY-2 -> FY) yields a 2-year CAGR.
    def hist(label):
        row = HIST_ROW[label]
        return [
            _get(cap, row, capiq_layout.COL_FY_M2),
            _get(cap, row, capiq_layout.COL_FY_M1),
            _get(cap, row, capiq_layout.COL_FY),
        ]

    revenue = hist("Revenue")
    cogs = hist("COGS")
    gp = hist("Gross Profit")
    opex = hist("Total Opex")
    da = hist("D&A")
    ebitda = hist("EBITDA")
    ebit = hist("EBIT")
    capex = hist("Capex")
    sbc = hist("SBC")
    dps = hist("DPS")

    def block(values):
        return {
            "FY-2": values[0],
            "FY-1": values[1],
            "FY":   values[2],
            "cagr_2y":   _cagr(values[0], values[2], 2),
            "yoy_latest": _yoy(values[1], values[2]),
        }

    historicals = {
        "revenue":      block(revenue),
        "cogs":         block(cogs),
        "gross_profit": block(gp),
        "total_opex":   block(opex),
        "d_and_a":      block(da),
        "ebitda":       block(ebitda),
        "ebit":         block(ebit),
        "capex":        block(capex),
        "sbc":          block(sbc),
        "dps":          block(dps),
    }

    def ratio_block(num, den):
        vals = [_safe_div(n, d) for n, d in zip(num, den)]
        return {
            "FY-2": vals[0], "FY-1": vals[1], "FY": vals[2],
            "avg_3y": _avg(vals),
        }

    ratios = {
        "gross_margin":   ratio_block(gp, revenue),
        "ebitda_margin": ratio_block(ebitda, revenue),
        "ebit_margin":   ratio_block(ebit, revenue),
        "opex_pct_rev":  ratio_block(opex, revenue),
        "capex_pct_rev": ratio_block(capex, revenue),
        "da_pct_capex":  ratio_block(da, capex),
        "sbc_pct_rev":   ratio_block(sbc, revenue),
    }

    # Current state — Section A (metadata) and Section B (current-state) values
    # all live in column F. Effective tax rate is no longer fetched.
    col_cur = capiq_layout.COL_CURRENT
    current_state = {
        "company_name":          _get(cap, META_ROW["Company Name"], col_cur),
        "sector":                _get(cap, META_ROW["Sector"], col_cur),
        "currency":              _get(cap, META_ROW["Currency"], col_cur),
        "filing_status":         _get(cap, META_ROW["Filing Status"], col_cur),
        "price":                 _get(cap, CUR_ROW["Current Price"], col_cur),
        "diluted_shares_mm":     _get(cap, CUR_ROW["Diluted Shares Out"], col_cur),
        "market_cap_mm":         _get(cap, CUR_ROW["Market Cap"], col_cur),
        "cash_mm":               _get(cap, CUR_ROW["Cash & Equivalents"], col_cur),
        "st_investments_mm":     _get(cap, CUR_ROW["ST Investments"], col_cur),
        "debt_mm":               _get(cap, CUR_ROW["Total Debt"], col_cur),
        "preferred_equity_mm":   _get(cap, CUR_ROW["Preferred Equity"], col_cur),
        "minority_interest_mm":  _get(cap, CUR_ROW["Minority Interest"], col_cur),
        "equity_investments_mm": _get(cap, CUR_ROW["Equity Investments"], col_cur),
        "marketable_securities_mm": _get(cap, CUR_ROW["Marketable Securities"], col_cur),
        "enterprise_value_mm":   _get(cap, CUR_ROW["Enterprise Value"], col_cur),
    }
    debt = current_state.get("debt_mm")
    cash = current_state.get("cash_mm")
    if isinstance(debt, (int, float)) and isinstance(cash, (int, float)):
        current_state["net_debt_mm"] = debt - cash
    # Annual DPS = most recent (FY) DPS, fall back to FY-1 if FY is blank.
    annual_dps = dps[2] if isinstance(dps[2], (int, float)) else dps[1]
    if isinstance(annual_dps, (int, float)):
        current_state["annual_dps"] = annual_dps

    return {
        "ticker": ticker,
        "company_name": current_state.get("company_name"),
        "sector": current_state.get("sector"),
        "currency": current_state.get("currency"),
        "fetch_timestamp": _get(cap, capiq_layout.ROW_FETCHER_DATE, 3),  # col C, written by fetch_capiq
        "historicals": historicals,
        "ratios": ratios,
        "current_state": current_state,
    }


def _fmt_pct(v): return f"{v*100:.1f}%" if isinstance(v, (int, float)) else "—"
def _fmt_num(v): return f"{v:,.0f}" if isinstance(v, (int, float)) else "—"
def _fmt_money(v): return f"${v:,.2f}" if isinstance(v, (int, float)) else "—"


def to_markdown(data: dict) -> str:
    h = data["historicals"]
    r = data["ratios"]
    cs = data["current_state"]
    lines = []
    lines.append(f"# {data['ticker']} — Historicals Brief")
    lines.append("")
    lines.append(f"**Company:** {data.get('company_name')}  ")
    lines.append(f"**Sector:** {data.get('sector')}  ")
    lines.append(f"**Currency:** {data.get('currency')}  ")
    lines.append(f"**Fetched:** {data.get('fetch_timestamp')}")
    lines.append("")
    lines.append("## Historicals")
    lines.append("")
    lines.append("| Metric | FY-2 | FY-1 | FY | 2Y CAGR | YoY latest |")
    lines.append("|---|---:|---:|---:|---:|---:|")
    for label, key in [
        ("Revenue", "revenue"), ("Gross Profit", "gross_profit"),
        ("Total OpEx", "total_opex"), ("EBITDA", "ebitda"),
        ("EBIT", "ebit"), ("D&A", "d_and_a"),
        ("CapEx", "capex"), ("SBC", "sbc"), ("DPS", "dps"),
    ]:
        b = h[key]
        lines.append(
            f"| {label} | {_fmt_num(b['FY-2'])} | {_fmt_num(b['FY-1'])} | {_fmt_num(b['FY'])} | "
            f"{_fmt_pct(b['cagr_2y'])} | {_fmt_pct(b['yoy_latest'])} |"
        )
    lines.append("")
    lines.append("## Ratios")
    lines.append("")
    lines.append("| Ratio | FY-2 | FY-1 | FY | 3Y avg |")
    lines.append("|---|---:|---:|---:|---:|")
    for label, key in [
        ("Gross Margin", "gross_margin"),
        ("EBITDA Margin", "ebitda_margin"),
        ("EBIT Margin", "ebit_margin"),
        ("OpEx % of Revenue", "opex_pct_rev"),
        ("CapEx % of Revenue", "capex_pct_rev"),
        ("D&A % of CapEx", "da_pct_capex"),
        ("SBC % of Revenue", "sbc_pct_rev"),
    ]:
        b = r[key]
        lines.append(
            f"| {label} | {_fmt_pct(b['FY-2'])} | {_fmt_pct(b['FY-1'])} | {_fmt_pct(b['FY'])} | "
            f"{_fmt_pct(b['avg_3y'])} |"
        )
    lines.append("")
    lines.append("## Current State")
    lines.append("")
    lines.append(f"- Price: {_fmt_money(cs.get('price'))}")
    lines.append(f"- Diluted shares (M): {_fmt_num(cs.get('diluted_shares_mm'))}")
    lines.append(f"- Market cap (M): {_fmt_num(cs.get('market_cap_mm'))}")
    lines.append(f"- Cash (M): {_fmt_num(cs.get('cash_mm'))}")
    lines.append(f"- Debt (M): {_fmt_num(cs.get('debt_mm'))}")
    lines.append(f"- Net debt (M): {_fmt_num(cs.get('net_debt_mm'))}")
    lines.append(f"- Enterprise value (M): {_fmt_num(cs.get('enterprise_value_mm'))}")
    lines.append(f"- Annual DPS: {_fmt_money(cs.get('annual_dps'))}")
    return "\n".join(lines)


def main(argv=None):
    parser = argparse.ArgumentParser(description="Extract historicals brief from _CapIQ_Data.")
    parser.add_argument("ticker")
    parser.add_argument("--format", choices=["json", "markdown"], default="json")
    parser.add_argument("--output", default=None,
                        help="Write to file instead of stdout.")
    args = parser.parse_args(argv)
    ticker = args.ticker.strip().upper()
    data = extract(ticker)
    if args.format == "json":
        out = json.dumps(data, indent=2, default=str)
    else:
        out = to_markdown(data)
    if args.output:
        Path(args.output).write_text(out, encoding="utf-8")
        print(f"Wrote {args.output}")
    else:
        sys.stdout.write(out + "\n")


if __name__ == "__main__":
    main()
